"""
views.py - Archivo principal de vistas de la aplicación
Contiene todas las vistas para manejar las diferentes funcionalidades del sistema
"""

# Importaciones de Django
from django.shortcuts import get_object_or_404, render, redirect
from django.views.generic import ListView, TemplateView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView  # Vistas genéricas básicas y para formularios
from django.urls import reverse_lazy, reverse  # Manejo de URLs
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash  # Funciones de autenticación
from django.contrib.auth.views import LoginView, PasswordChangeView  # Vistas de autenticación
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin  # Mixins de autenticación
from django.contrib.auth.decorators import login_required # Importación para el decorador
from django.contrib.auth.models import Group, User  # Modelos de autenticación
from django.contrib import messages  # Sistema de mensajes
from django.conf import settings  # Configuraciones
from django.http import HttpResponse, HttpResponseRedirect, Http404  # Respuestas HTTP
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger  # Paginación
from django.db.models import F, Value, CharField, Q, Case, When  # Operaciones de base de datos
from django.db.models.functions import Concat
from django.utils.timezone import now
from django.utils import timezone
from django.db import transaction, IntegrityError
from django.apps import apps  # Aplicaciones de Django
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.db.models import Q
from django.contrib.contenttypes.models import ContentType # Para obtener el modelo dinámicamente
from django.contrib.contenttypes.models import ContentType # Para obtener el modelo dinámicamente
from django.db import models
from .models import Profile
from decimal import Decimal
from uuid import UUID

from datetime import datetime
from django.utils.timezone import make_aware

from django.views.generic.detail import DetailView
# Importaciones de terceros
import pandas as pd
import numpy as np # Importar numpy
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import logging
from .opciones import opciones_estado_activo, opciones_activos, opciones_marca_all_in_one, opciones_marca_notebook, opciones_marca_mini_pc, opciones_marca_proyector


# Importaciones locales - Modelos
from accounts.models import Profile  # Modelo de perfil de usuario
from .models import (
    AllInOne, AllInOneAdmins, Notebook, MiniPC,
    Proyectores, BodegaADR, Azotea, Profile, Eliminados, HistorialCambios,
    Monitor, Audio, Tablet  # Nuevos modelos agregados
)

# Importaciones locales - Formularios
from .forms import (
    LoginForm, UserCreationForm, ProfileForm, UserForm, RegisterUserForm,
    AllInOneForm, AllInOneAdminsForm, NotebooksForm, MiniPCForm,
    ProyectoresForm, BodegaADRForm, AzoteaForm,
    UploadExcelForm, MonitorForm, AudioForm, TabletForm # Nuevos formularios agregados
)
from django.forms.models import model_to_dict


# Importaciones locales - Funciones y decoradores
from .funciones import plural_singular, filtrar_y_paginar
from .decorators import add_group_name_to_context, get_group_and_color


from django.core.mail import send_mail

from django.core.mail import send_mail, EmailMessage # Importar EmailMessage
from .forms import ProfileImageForm
import os # Importar os
import tempfile # Importar tempfile
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import PasswordChangeView
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
from .forms import LoginForm,UserUpdateForm
from adr.utils import enviar_notificacion_asunto

#logica para que el usuario pueda editar su foto
@login_required
def my_profile(request):
    user = request.user
    profile = user.profile

    uform = UserUpdateForm(instance=user)
    pform = ProfileImageForm(instance=profile)

    if request.method == "POST":
        # 1) Guardar datos
        if "save_user" in request.POST:
            uform = UserUpdateForm(request.POST, instance=user)
            if uform.is_valid():
                uform.save()
                messages.success(request, "Datos actualizados correctamente.")
                return redirect("my_profile")
            else:
                messages.error(request, "Revisa los campos del formulario.")

        # 2) Guardar foto
        elif "save_photo" in request.POST:
            pform = ProfileImageForm(request.POST, request.FILES, instance=profile)
            if pform.is_valid():
                pform.save()
                messages.success(request, "Tu foto de perfil fue actualizada.")
                return redirect("my_profile")
            else:
                messages.error(request, "No se pudo actualizar la foto. Revisa el archivo.")

        # 3) **ELIMINAR foto**  <-- Faltaba esto
        elif "delete_image" in request.POST:
            if profile.image:
                # borra el archivo del disco y limpia el campo
                profile.image.delete(save=False)
                profile.image = None
                profile.save()
                messages.success(request, "La foto de perfil fue eliminada.")
            else:
                messages.info(request, "No tenías una foto subida.")
            return redirect("my_profile")

    group = request.user.groups.first().name if request.user.groups.exists() else "Invitado"

    return render(
        request,
        "profiles/my_profile.html",
        {
            "user_form": uform,
            "photo_form": pform,
            "user_profile": user,
            "group_name_singular": group,
        },
    )
from io import BytesIO
from uuid import uuid4
from PIL import Image, ImageOps
from django.core.files.base import ContentFile
from django.contrib.auth.views import PasswordResetView

def make_avatar_square(django_file, size=512, fmt="WEBP", quality=86):
    """
    - Corrige orientación EXIF
    - Recorte centrado a cuadrado
    - Redimensiona con LANCZOS
    - Exporta a WEBP (o JPEG)
    """
    img = Image.open(django_file)
    img = ImageOps.exif_transpose(img)     # corrige orientación
    img = img.convert("RGB")

    w, h = img.size
    side = min(w, h)
    left = (w - side) // 2
    top  = (h - side) // 2
    img = img.crop((left, top, left + side, top + side))
    img = img.resize((size, size), Image.LANCZOS)

    buf = BytesIO()
    if fmt.upper() == "WEBP":
        img.save(buf, "WEBP", quality=quality, method=6)
        ext = "webp"
    else:
        img.save(buf, "JPEG", quality=quality, optimize=True, progressive=True)
        ext = "jpg"

    name = f"avatar_{uuid4().hex}.{ext}"
    return ContentFile(buf.getvalue(), name=name)
class UserPasswordChangeView(LoginRequiredMixin, SuccessMessageMixin, PasswordChangeView):
    template_name = "profiles/password_change.html"
    success_url = reverse_lazy("my_profile")      # <-- vuelve al perfil
    success_message = "Tu contraseña se cambió correctamente."
# Vista para descargar archivos Excel
class CustomPasswordResetView(PasswordResetView):
    email_template_name = 'registration/password_reset_email.txt'
    def dispatch(self, request, *args, **kwargs):
        messages.success(request, "Tu contraseña fue cambiada correctamente.")
        return redirect(reverse_lazy('inicio'))  # Cambia 'inicio' por el nombre de tu URL de inicio
@add_group_name_to_context
class DescargarExcelView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista para descargar datos en formato Excel"""

    # Revisa los grupos del usuario antes de permitir el acceso
    def test_func(self):
        """Solo ADR o Operadores ADR pueden descargar el archivo"""
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']).exists()

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get(self, request, *args, **kwargs):
        model_name = kwargs.get('model_name')
        fecha_actual = now().strftime('%d-%m-%Y')

        # Diccionario que mapea nombres de modelos a sus clases y nombres de archivo
        model_mapping = {
            'allinone': (AllInOne, f'AllInOne_{fecha_actual}.xlsx'),
            'allinoneadmin': (AllInOneAdmins, f'AllInOneAdmins_{fecha_actual}.xlsx'),
            'notebook': (Notebook, f'Notebooks_{fecha_actual}.xlsx'),
            'minipc': (MiniPC, f'MiniPCs_{fecha_actual}.xlsx'),
            'proyector': (Proyectores, f'Proyectores_{fecha_actual}.xlsx'),
            'bodegaadr': (BodegaADR, f'BodegaADR_{fecha_actual}.xlsx'),
            'azotea': (Azotea, f'Azotea_{fecha_actual}.xlsx'),
            'eliminados': (Eliminados, f'Eliminados_{fecha_actual}.xlsx'),
            'historialcambios': (HistorialCambios, f'HistorialCambios_{fecha_actual}.xlsx'),
            'monitor': (Monitor, f'Monitores_{fecha_actual}.xlsx'),
            'audio': (Audio, f'Audio_{fecha_actual}.xlsx'),
            'tablet': (Tablet, f'Tablets_{fecha_actual}.xlsx')
        }

        # Validación del modelo
        if model_name not in model_mapping:
            return HttpResponse(status=404)

        # Obtiene el modelo y nombre de archivo correspondiente
        model_class, filename = model_mapping[model_name]

        # Verifica si hay un filtro de búsqueda en los parámetros de la URL
        search_query = request.GET.get('search', '').strip()
        if search_query:
            queryset = model_class.objects.filter(
                Q(activo__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(ubicacion__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query)
            )
        else:
            queryset = model_class.objects.all()

        # Crea el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = model_name.capitalize()

        # Agrega encabezados basados en los campos del modelo
        columns = [field.name for field in model_class._meta.fields]
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            ws[f"{column_letter}1"] = column_title.capitalize()

        # Rellena los datos de cada objeto del queryset en la hoja
        for row_num, obj in enumerate(queryset, 2):  # Empieza desde la fila 2
            for col_num, field_name in enumerate(columns, 1):
                column_letter = get_column_letter(col_num)
                field_value = getattr(obj, field_name)
                ws[f"{column_letter}{row_num}"] = str(field_value) if field_value is not None else ''

        # Configura la respuesta HTTP para la descarga
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Guarda el archivo en la respuesta
        wb.save(response)
        return response


# -------- VISTAS DE AUTENTICACIÓN Y PERFILES --------

@add_group_name_to_context
class Login(LoginView):
    """Vista de inicio de sesión"""
    template_name = 'index.html'
    form_class = LoginForm
    success_url = reverse_lazy('home')

@add_group_name_to_context
class IndexView(TemplateView):
    """Vista de la página principal"""
    template_name = 'index.html'

@add_group_name_to_context
class HomeView(LoginRequiredMixin, TemplateView):
    """Vista del dashboard principal (requiere autenticación)"""
    template_name = 'home.html'

@add_group_name_to_context
class AllInOneSelectionView(LoginRequiredMixin, TemplateView):
    """Vista para seleccionar el tipo de All In One"""
    template_name = 'all_in_one_selection.html'
@add_group_name_to_context
class ErrorView(TemplateView):
    """Vista para mostrar errores"""
    template_name = 'error.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['error_image'] = os.path.join(settings.MEDIA_ROOT, 'error.png')
        return context

@add_group_name_to_context
class AddUserView(UserPassesTestMixin, LoginRequiredMixin, CreateView):
    """Vista para agregar nuevos usuarios (solo ADR)"""
    model = User
    form_class = RegisterUserForm
    template_name = 'profiles/add_user.html'
    success_url = reverse_lazy('profile_list')

    def test_func(self):
        """Verifica que el usuario sea ADR"""
        first_group = self.request.user.groups.first()
        return bool(first_group and first_group.name == 'ADR')

    def handle_no_permission(self):
        """Redirecciona a error si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega grupos al contexto"""
        context = super().get_context_data(**kwargs)
        context['singular_groups'] = Group.objects.values_list('name', 'id')
        return context

    def form_valid(self, form):
        """Crea el usuario, marca 'create_by_adr=True', asigna grupo y notifica"""
        try:
            group_id = self.request.POST.get('group')
            if not group_id:
                messages.error(self.request, 'Debe seleccionar un grupo para el usuario.')
                return self.form_invalid(form)

            try:
                group = Group.objects.get(id=group_id)
            except Group.DoesNotExist:
                messages.error(self.request, 'El grupo seleccionado no existe.')
                return self.form_invalid(form)

            with transaction.atomic():
                # 1) Crear usuario
                user = form.save(commit=False)
                user.first_name = form.cleaned_data.get('first_name', '')
                user.last_name  = form.cleaned_data.get('last_name', '')
                # Asegura setear una contraseña válida (viene de tu form password1)
                raw_password = form.cleaned_data.get('password1', '')
                user.set_password(raw_password)

                # (opcional) regla para is_staff según grupos
                # ajusta a tu lógica real en lugar del id fijo '2'
                if group.name in ['ADR', 'Operadores ADR']:
                    user.is_staff = True

                user.save()

                # 2) Asignar grupo
                user.groups.clear()
                user.groups.add(group)

                # 3) Crear/actualizar Profile y marcar para cambio de password
                profile, _ = Profile.objects.get_or_create(user=user)
                profile.create_by_adr = True   # ← clave para forzar cambio al primer login
                profile.save(update_fields=['create_by_adr'])

            # 4) Notificación por correo (opcional)
            try:
                accion = "Nuevo Usuario Agregado"
                mensaje = f"""
El usuario {self.request.user.get_full_name()} ha agregado un nuevo usuario al sistema.

Acción: {accion}
Nombre Completo del Usuario Agregado: {user.first_name} {user.last_name}
Nombre de Usuario: {user.username}
Grupo Asignado: {group.name}
"""
                enviar_notificacion_asunto(
                    asunto="Nuevo Usuario Registrado en el Sistema",
                    mensaje=mensaje,
                    destinatarios=getattr(settings, 'EMAIL_RECIPIENTS', [])
                )
            except Exception as e:
                # No detengas la creación por fallo de correo
                messages.warning(self.request, f'Usuario creado, pero falló el envío de correo: {str(e)}')

            messages.success(self.request, 'Usuario creado exitosamente.')
            return redirect(self.success_url)

        except Exception as e:
            messages.error(self.request, f'Error al crear el usuario: {str(e)}')
            return self.form_invalid(form)

# -------- VISTAS DE PERFILES --------

@add_group_name_to_context
class ProfileListView(LoginRequiredMixin, ListView):
    """Vista para listar los perfiles de usuarios"""
    model = Profile
    template_name = 'profiles/profile_list.html'
    context_object_name = 'profiles'
    paginate_by = 25

    def get_queryset(self):
        """
        Obtiene y configura el queryset de perfiles
        Añade el nombre del grupo y ordena por grupo y nombre de usuario
        """
        queryset = super().get_queryset()
        queryset = queryset.annotate(
            group_name=F('user__groups__name')
        ).order_by('-group_name', 'user__username')
        return queryset

    def get_context_data(self, **kwargs):
        """
        Añade datos adicionales al contexto de la plantilla
        - Información del grupo del usuario actual
        - Lista de perfiles con sus grupos en formato singular
        """
        context = super().get_context_data(**kwargs)
        user = self.request.user
        group_id, group_name, group_name_singular, color = get_group_and_color(user)
        context['group_name'] = group_name
        context['group_name_singular'] = group_name_singular
        context['color'] = color
        
        profiles_with_singular_groups = []
        for profile in context['profiles']:
            singular_groups = [plural_singular(group.name) for group in profile.user.groups.all()]
            profiles_with_singular_groups.append({
                'profile': profile,
                'singular_groups': singular_groups
            })

        context['profiles_with_singular_groups'] = profiles_with_singular_groups
        return context

@add_group_name_to_context
class ProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = Profile
    template_name = 'profiles/profile_edit.html'
    context_object_name = 'user_profile'
    form_class = ProfileForm  # Mantén ProfileForm para el perfil

    def get_object(self):
        """Obtiene el perfil a editar usando el pk del perfil"""
        return get_object_or_404(Profile, user__pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        """
        Prepara el contexto para la edición del perfil
        - Incluye formularios de usuario y perfil
        - Añade información de grupos
        """
        context = super().get_context_data(**kwargs)
        profile = self.get_object()
        user = profile.user
        context['user_profile'] = user
        context['profile_form'] = ProfileForm(instance=profile)
        context['user_form'] = UserForm(instance=user)
        context['singular_groups'] = Group.objects.values_list('name', 'id')
        context['group_id_user'] = user.groups.values_list('id', flat=True).first()
        return context

    def post(self, request, *args, **kwargs):
        profile = self.get_object()
        user = profile.user

        # --- NUEVO: eliminar solo la foto de perfil (sin tocar otros datos) ---
        if 'delete_image' in request.POST:
            profile.clear_image()
            messages.success(request, 'La foto de perfil fue eliminada y se restauró la imagen por defecto.')
            return redirect(request.path)

        # --- Flujo normal de actualización ---
        user_form = UserForm(request.POST, instance=user)
        profile_form = ProfileForm(request.POST, request.FILES, instance=profile)

        # Obtener las nuevas contraseñas
        new_password1 = request.POST.get("new_password1")
        new_password2 = request.POST.get("new_password2")

        if user_form.is_valid() and profile_form.is_valid():
            # Guardar usuario y perfil
            user_form.save()
            profile_form.save()

            # Asignar grupo
            group_id = request.POST.get('group')
            grupo_asignado = "No asignado"
            if group_id:
                new_group = Group.objects.get(id=group_id)
                user.groups.clear()
                user.groups.add(new_group)
                grupo_asignado = new_group.name

            # Cambio de contraseña (opcional)
            password_cambiada = False
            if new_password1 and new_password1 == new_password2:
                user.set_password(new_password1)
                user.save()
                password_cambiada = True
                messages.success(request, 'Contraseña actualizada exitosamente')
            elif new_password1 or new_password2:
                messages.error(request, 'Las contraseñas no coinciden. Por favor, intente nuevamente.')

            # Notificación por correo
            accion = "Actualización de Perfil"
            mensaje = f"""
            El usuario {request.user.get_full_name()} ha actualizado el perfil de usuario.

            Acción: {accion}
            Nombre de Usuario Actualizado: {user.username}
            Nombre Completo: {user.first_name} {user.last_name}
            Grupo Asignado: {grupo_asignado}
            Contraseña: {"Cambiada" if password_cambiada else "No Cambiada"}
            """
            try:
                enviar_notificacion_asunto(
                    asunto="Actualización de Perfil de Usuario",
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )
            except Exception as e:
                print(f"Error al enviar el correo de notificación: {str(e)}")
                messages.error(request, 'Error al enviar el correo de notificación.')

            messages.success(request, 'Usuario editado exitosamente')
            return redirect('profile_list')

        # Errores de validación
        context = self.get_context_data()
        context['user_form'] = user_form
        context['profile_form'] = profile_form
        return render(request, 'profiles/profile_edit.html', context)

    def get_success_url(self):
        """Redirige después de guardar"""
        return reverse_lazy('profile_list')


# --- Helper seguro para enviar correos (no rompe si falla) ---
def _enviar_notificacion(asunto: str, mensaje: str, destinatarios: list[str] | tuple[str, ...] | None):
    """
    Envía un correo simple. Si no hay destinatarios o falla, no levanta excepción.
    Usa DEFAULT_FROM_EMAIL si está definido.
    """
    try:
        if not destinatarios:
            return  # sin destinatarios, no envía
        from_email = getattr(settings, "DEFAULT_FROM_EMAIL", None)
        send_mail(
            subject=asunto,
            message=mensaje,
            from_email=from_email,
            recipient_list=list(destinatarios),
            fail_silently=True,  # importantísimo para no romper el flujo
        )
    except Exception:
        # No hacemos nada: el borrado no debe fallar por el correo
        pass


class ProfileDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = User
    success_url = reverse_lazy('profile_list')
    template_name = 'profiles/profile_confirm_delete.html'

    def test_func(self):
        """Solo ADR puede eliminar perfiles"""
        return self.request.user.groups.filter(name='ADR').exists()

    def handle_no_permission(self):
        messages.error(self.request, 'No tiene permisos para esta acción')
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.groups.exists():
            group_name = self.request.user.groups.first().name
            context['group_name_singular'] = group_name.replace('es ADR', ' ADR').replace('s ADR', ' ADR')
        return context

    def delete(self, request, *args, **kwargs):
        """Procesa la eliminación de un usuario y envía notificación (sin romper si el correo falla)"""
        try:
            self.object = self.get_object()

            # --- Evitar auto-eliminación (opcional, recomendado) ---
            if self.object == request.user:
                messages.error(request, 'No puedes eliminar tu propia cuenta.')
                return redirect(self.success_url)

            nombre_usuario = self.object.username
            nombre_completo = f"{self.object.first_name} {self.object.last_name}".strip()
            grupo = self.object.groups.first().name if self.object.groups.exists() else "Sin grupo asignado"

            # Eliminar el usuario
            self.object.delete()

            # Preparar y enviar notificación (si hay destinatarios configurados)
            accion = "Eliminación de Perfil"
            mensaje = (
                f"El usuario {request.user.get_full_name() or request.user.username} ha eliminado "
                f"el siguiente perfil de usuario:\n\n"
                f"Acción: {accion}\n"
                f"Nombre de Usuario Eliminado: {nombre_usuario}\n"
                f"Nombre Completo: {nombre_completo or '-'}\n"
                f"Grupo Asignado: {grupo}\n"
            )

            # Puedes usar una lista en settings: EMAIL_RECIPIENTS = ["soporte@tudominio.cl", ...]
            destinatarios = getattr(settings, "EMAIL_RECIPIENTS", None)
            _enviar_notificacion(
                asunto="Eliminación de Perfil de Usuario",
                mensaje=mensaje,
                destinatarios=destinatarios,
            )

            messages.success(self.request, f'Usuario {nombre_usuario} eliminado exitosamente')
            return HttpResponseRedirect(self.get_success_url())

        except Exception as e:
            messages.error(self.request, f'Error al eliminar usuario: {e}')
            return redirect('profile_list')

    def post(self, request, *args, **kwargs):
        # El botón del template hace POST, así que delegamos en delete()
        return self.delete(request, *args, **kwargs)





@add_group_name_to_context
class ProfilePasswordChangeView(PasswordChangeView):
    """Vista para cambio de contraseña de perfil"""
    template_name = 'profiles/change_password.html'
    success_url = reverse_lazy('home')

    def get_context_data(self, **kwargs):
        """Añade estado de cambio de contraseña al contexto"""
        context = super().get_context_data(**kwargs)
        context['password_changed'] = self.request.session.get('password_changed', False)
        return context
    
    def form_valid(self, form):
        """
        Procesa el cambio de contraseña exitoso
        - Actualiza el estado del perfil
        - Establece mensajes de éxito
        - Actualiza la sesión
        """
        profile = Profile.objects.get(user=self.request.user)
        profile.create_by_adr = False
        profile.save()

        messages.success(self.request, 'Contraseña cambiada exitosamente')
        update_session_auth_hash(self.request, form.user)
        self.request.session['profile_password_changed'] = True
        return super().form_valid(form)
    
    def form_invalid(self, form):
        """Manejo de formulario inválido con mensaje de error"""
        messages.error(self.request, 'Las contraseñas no coinciden o no cumple el estándar de seguridad')
        return super().form_invalid(form)

@add_group_name_to_context
class CustomLoginView(LoginView):
    def form_invalid(self, form):
        messages.error(self.request, 'Usuario o contraseña incorrectos. Por favor, intente nuevamente.')
        return super().form_invalid(form)

    def form_valid(self, form):
        user = form.get_user()
        profile = getattr(user, "profile", None)

        if profile and profile.create_by_adr:
            messages.warning(self.request, 'Bienvenido, debes cambiar tu contraseña ahora.')
            return HttpResponseRedirect(reverse_lazy('profile_password_change'))

        resp = super().form_valid(form)
        messages.success(self.request, 'Inicio de sesión exitoso.')
        return resp

# -------- VISTAS DE ALL IN ONE --------

@add_group_name_to_context
class AllInOneView(LoginRequiredMixin, ListView):
    """Vista para listar todos los equipos All In One"""
    model = AllInOne
    template_name = 'modulos/all_in_one.html'
    context_object_name = 'all_in_ones'
    paginate_by = 25  # Paginación de 25 items por página
    ordering = ['ubicacion', '-fecha_creacion']  # Ordenado por ubicación y fecha

    def get_queryset(self):
        """Obtiene y filtra la lista según búsqueda"""
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        search_by_pk = self.request.GET.get('search_by_pk', 'false').lower() == 'true'

        # Validar si el parámetro es búsqueda por PK y si el input es numérico
        if search_by_pk:
            if search_query.isdigit():  # Solo buscar por PK si es un número
                return queryset.filter(pk=int(search_query))
            else:
                # Si no es numérico, devolver queryset vacío
                return queryset.none()
       
        if search_query:
            # Búsqueda en múltiples campos
            queryset = queryset.filter(
                Q(activo__icontains=search_query) |
                Q(estado__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(ubicacion__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query) |
                Q(fecha_creacion__icontains=search_query)
            )
       
        return queryset.select_related('creado_por').order_by('ubicacion', '-fecha_creacion')

    def get_context_data(self, **kwargs):
        """Agrega el término de búsqueda al contexto"""
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '').strip()
        return context

@add_group_name_to_context
class Add_AllInOneView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo All In One"""
    model = AllInOne
    template_name = './modulos/add_all_in_one.html'
    form_class = AllInOneForm
    success_url = reverse_lazy('all_in_one')

    def test_func(self):
        """Verifica permisos: solo ADR, Operadores y Auxiliares pueden agregar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega lista de All In One al contexto"""
        context = super().get_context_data(**kwargs)
        context['all_in_ones'] = AllInOne.objects.all()
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con validaciones adicionales y envía notificación"""
        try:
            if form.is_valid():
                # Validación del número de serie
                n_serie = form.cleaned_data.get('n_serie')
                if not n_serie:
                    form.add_error('n_serie', 'El número de serie es requerido')
                    return self.form_invalid(form)

                if AllInOne.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este Número de Serie ya existe')
                    return self.form_invalid(form)

                # Validación de campos obligatorios
                required_fields = {
                    'modelo': 'El modelo es requerido',
                    'unive': 'El UNIVE es requerido',
                    'netbios': 'El NetBios es requerido',
                }

                for field, error_message in required_fields.items():
                    if not form.cleaned_data.get(field):
                        form.add_error(field, error_message)
                        return self.form_invalid(form)

                # Validación del formato BDO (permite 0)
                bdo = form.cleaned_data.get('bdo')
                if bdo is not None: # Solo validar si se proporcionó un valor
                    try:
                        bdo_int = int(bdo)
                        if bdo_int < 0:
                            form.add_error('bdo', 'El campo BDO no puede ser un número negativo.')
                            return self.form_invalid(form)
                    except ValueError:
                        form.add_error('bdo', 'El campo BDO solo debe contener números.')
                        return self.form_invalid(form)

                # Guarda con el usuario actual
                form.instance.creado_por = self.request.user
                instance = form.save()

                # Enviar notificación por correo
                accion = "Registro Agregado"
                modelo = "All In One"
                user = self.request.user
                user_group = user.groups.first().name if user.groups.exists() else "Sin grupo asignado"
                mensaje = f"""
                El usuario {user.get_full_name()} (Grupo: {user_group}) ha realizado la siguiente acción:
                Acción: {accion}
                Modelo: {modelo}
                Datos:
                - Estado: {instance.estado}
                - Activo: {modelo}
                - Marca: {instance.marca}
                - Modelo: {instance.modelo}
                - N° Serie: {instance.n_serie}
                - UNIVE: {instance.unive}
                - NetBIOS: {getattr(instance, 'netbios', '')}
                - BDO: {getattr(instance, 'bdo', '')}
                - Ubicación: {getattr(instance, 'ubicacion', '')}

                """

                enviar_notificacion_asunto(
                    asunto="Registro Agregado en All In One",
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )

                messages.success(self.request, 'All In One se ha guardado correctamente.')
                return super().form_valid(form)

            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)


@add_group_name_to_context
class Edit_AllInOneView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar All In One existente"""
    model = AllInOne
    template_name = 'modulos/edit_all_in_one.html'
    form_class = AllInOneForm
    success_url = reverse_lazy('all_in_one')

    def test_func(self):
        """Solo ADR y Operadores pueden editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['opciones_marca_all_in_one'] = opciones_marca_all_in_one
        # Añadir el nombre del modelo para el botón de eliminar
        context['model_name'] = self.model._meta.model_name
        return context
    
    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación"""
        try:
            # Validación de campos obligatorios
            required_fields = {
                'modelo': 'El modelo es requerido',
                'unive': 'El UNIVE es requerido',
                'netbios': 'El NetBios es requerido',
                # 'bdo': 'El BDO es requerido'  # Añadimos BDO como campo requerido
            }
            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación del formato BDO
            bdo = form.cleaned_data.get('bdo')
            try:
                int(bdo)
            except ValueError:
                form.add_error('bdo', 'El campo BDO solo debe contener números')
                return self.form_invalid(form)

            # Mantiene el creador original
            # Obtener el objeto original antes de guardar para comparar los cambios
            # Obtener el objeto original antes de guardar para comparar los cambios
            original_instance = AllInOne.objects.get(pk=self.object.pk)

            # Obtener el objeto original antes de guardar para comparar los cambios
            original_instance = AllInOne.objects.get(pk=self.object.pk)

            # Actualizar la instancia manualmente con los datos limpios del formulario
            instance = self.object # Usar la instancia existente
            cambios_detectados = []
            modelo_modificado = self.model._meta.verbose_name
            usuario_modificador = self.request.user

            for field_name, new_value in form.cleaned_data.items():
                # Excluir campos que no queremos registrar o que se manejan automáticamente
                if field_name in ['creado_por', 'fecha_creacion', 'fecha_modificacion', 'activo']:
                    continue

                old_value = getattr(original_instance, field_name)

                # Convertir Decimal a string para comparación si es necesario
                if isinstance(old_value, Decimal):
                    old_value = str(old_value)
                if isinstance(new_value, Decimal):
                    new_value = str(new_value)

                # Convertir UUID a string para comparación
                if isinstance(old_value, UUID):
                    old_value = str(old_value)
                if isinstance(new_value, UUID):
                    new_value = str(new_value)

                # Manejar casos de None y valores vacíos para comparación
                if old_value is None:
                    old_value = ''
                if new_value is None:
                    new_value = ''

                # Comparar y actualizar solo si el valor ha cambiado, normalizando los valores para una comparación más robusta
                if str(old_value).strip().lower() != str(new_value).strip().lower():
                    setattr(instance, field_name, new_value) # Actualizar el atributo de la instancia
                    HistorialCambios.objects.create(
                        modelo=modelo_modificado,
                        objeto_id=instance.pk,
                        usuario=usuario_modificador,
                        campo_modificado=field_name,
                        valor_anterior=str(old_value),
                        valor_nuevo=str(new_value)
                    )
                    cambios_detectados.append(f"- {field_name.replace('_', ' ').capitalize()}: De '{old_value}' a '{new_value}'")

            # Guardar la instancia después de actualizar los campos
            instance.save()

            # Enviar notificación por correo solo si hay cambios detectados
            if cambios_detectados:
                accion = "Registro Modificado"
                user = self.request.user
                user_group = user.groups.first().name if user.groups.exists() else "Sin grupo asignado"
                  # 1) Prepara la cadena de cambios fuera del f-string
                detalle_cambios = "\n".join(cambios_detectados)

                # 2) Construye el mensaje sin backslashes dentro de las llaves
                mensaje = f"""
                El usuario {user.get_full_name()} (Grupo: {user_group}) ha realizado la siguiente acción:
                Acción: {accion}
                Modelo: {modelo_modificado}
                Datos Modificados:
                {detalle_cambios}
                """
                enviar_notificacion_asunto(
                    asunto=f"Registro Modificado en {modelo_modificado}",
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )
                messages.success(self.request, 'Registro All In One se ha modificado correctamente.')
            else:
                messages.info(self.request, 'No se detectaron cambios en el registro All In One.')

            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)




logger = logging.getLogger(__name__)

# Diccionario que mapea nombres de modelos con los modelos reales
MODELS_DICT = {
    'all_in_one': AllInOne,
    'all_in_one_admin': AllInOneAdmins,
    'mini_pc': MiniPC,
    'notebook': Notebook,
    'proyector': Proyectores,
    'bodegaadr': BodegaADR,
    'azotea': Azotea,
    'monitor': Monitor,
    'audio': Audio,
    'tablet': Tablet,
}


@add_group_name_to_context
class EliminadosListView(LoginRequiredMixin, ListView):
    model = Eliminados
    template_name = 'modulos/eliminados.html'
    context_object_name = 'eliminados'
    paginate_by = 25  # Número de objetos por página
    
    def test_func(self):
        """Verifica permisos: solo ADR y Operadores pueden agregar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener el valor del filtro de búsqueda desde la solicitud GET
        search_query = self.request.GET.get('search', '').strip()

        # Obtener el queryset y aplicar el filtro de búsqueda
        queryset = Eliminados.objects.all()
        if search_query:
            queryset = queryset.filter(    
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(estado__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(ubicacion__icontains=search_query) |
                Q(eliminado_por__first_name__icontains=search_query) |
                Q(eliminado_por__last_name__icontains=search_query) |
                Q(fecha_eliminacion__icontains=search_query)
            )

        # Ordenar el queryset
        queryset = queryset.order_by('-fecha_eliminacion')

        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            eliminados_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            eliminados_paginated = paginator.page(1)
        except EmptyPage:
            eliminados_paginated = paginator.page(paginator.num_pages)

        # Agregar los datos al contexto
        context['eliminados'] = eliminados_paginated
        context['search_query'] = search_query

        return context

@add_group_name_to_context
class ConfirmarRestauracionView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Eliminados
    template_name = 'confirmar_restauracion_allinone.html'
    context_object_name = 'registro_eliminado'

    def test_func(self):
        """Solo ADR y Operadores pueden restaurar registros"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def post(self, request, *args, **kwargs):
        """Restaurar el registro de Eliminados a su tabla original"""
        registro = self.get_object()
        # Ajustar el nombre del modelo eliminando espacios y poniéndolo en minúsculas
        model_name = registro.activo.lower().replace(' ', '')
        model = MODELS_DICT.get(model_name)

        if not model:
            messages.error(request, f'Modelo no encontrado para restaurar: {model_name}')
            return redirect('eliminados')

        try:
            # Obtener solo los campos que existen en el modelo destino
            campos_validos = {field.name for field in model._meta.get_fields()}
            datos_restauracion = {
                'activo': registro.activo,
                'modelo': registro.modelo,
                'n_serie': registro.n_serie,
                'unive': registro.unive,
                'bdo': registro.bdo,
                'estado': registro.estado,
                'marca': registro.marca,
                'creado_por': registro.eliminado_por,
                'fecha_creacion': registro.fecha_eliminacion
            }

            # Agregar los campos opcionales si están en el modelo destino
            if 'netbios' in campos_validos:
                datos_restauracion['netbios'] = registro.netbios
            if 'ubicacion' in campos_validos:
                datos_restauracion['ubicacion'] = registro.ubicacion

            # Crear una nueva instancia del modelo con los datos restaurados
            restored_instance = model.objects.create(**datos_restauracion)

            # Enviar notificación por correo
            accion = "Restauración de Registro Eliminado"
            modelo = registro.activo
            mensaje = f"""
            El usuario {request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos Restaurados:
            - Activo: {registro.activo}
            - Marca: {registro.marca}
            - Modelo: {registro.modelo}
            - N° Serie: {registro.n_serie}
            - UNIVE: {registro.unive}
            - BDO: {registro.bdo}
            - NetBIOS: {registro.netbios if 'netbios' in campos_validos else 'No aplica'}
            - Ubicación: {registro.ubicacion if 'ubicacion' in campos_validos else 'No aplica'}
            """

            enviar_notificacion_asunto(
                asunto="Restauración de Registro Eliminado",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            # Eliminar de la tabla Eliminados
            registro.delete()
            messages.success(request, 'El registro ha sido restaurado correctamente.')
            return redirect('eliminados')

        except Exception as e:
            messages.error(request, f'Error al restaurar el registro: {str(e)}')
            return redirect('eliminados')







# Configurar el logger
@add_group_name_to_context
class DeleteToEliminadosView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vista genérica para eliminar registros moviéndolos a la tabla de Eliminados"""
    success_url = reverse_lazy('eliminados')  # Redireccionamiento por defecto

    def test_func(self):
        """Solo ADR y Operadores puede eliminar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        return redirect('error')

    def post(self, request, *args, **kwargs):
        """Procesa la eliminación lógica moviendo el registro a la tabla de Eliminados"""
        model_name = kwargs.get('model_name')
        pk = kwargs.get('pk')

        # Validación de existencia de model_name y pk
        if not model_name or not pk:
            messages.error(request, 'Información del modelo o registro no proporcionada correctamente.')
            return redirect(self.success_url)

        # Normaliza el nombre del modelo para buscarlo en el diccionario
        model_name = model_name.lower()  # Asegurarse de que el nombre sea siempre minúscula para la comparación

        # Debugging: Print model_name and MODELS_DICT keys
        print(f"Attempting to delete model: {model_name}")
        print(f"Available models in MODELS_DICT: {MODELS_DICT.keys()}")

        model = MODELS_DICT.get(model_name)
        if not model:
            messages.error(request, f'Modelo no encontrado: {model_name}')
            return redirect(self.success_url)

        try:
            # Obtener el objeto que se quiere eliminar
            instance = get_object_or_404(model, pk=pk)

            # Uso de una transacción atómica para garantizar la consistencia de los datos
            with transaction.atomic():
                # Intentar guardar los datos en la tabla Eliminados
                eliminado_data = {
                    'activo': model_name.title(),
                    'modelo': instance.modelo,
                    'n_serie': instance.n_serie if instance.n_serie is not None else '', # Asigna '' si es None
                    'unive': instance.unive,
                    'bdo': instance.bdo,
                    'estado': instance.estado,
                    'marca': instance.marca,
                    'eliminado_por': request.user,
                    'fecha_eliminacion': timezone.now()
                }

                # Incluir netbios y ubicacion solo si existen en el modelo original
                if hasattr(instance, 'netbios'):
                    eliminado_data['netbios'] = getattr(instance, 'netbios', '')
                if hasattr(instance, 'ubicacion'):
                     eliminado_data['ubicacion'] = getattr(instance, 'ubicacion', '')

                print("DEBUG: Intentando crear registro en Eliminados con data:", eliminado_data) # Debug print
                eliminado = Eliminados.objects.create(**eliminado_data)

                # Si se creó correctamente en Eliminados, eliminar de la tabla original
                if eliminado:
                    instance.delete()  # Se elimina solo si se completó correctamente la creación en Eliminados
                    
                    # Enviar notificación por correo
                    accion = "Eliminación Lógica (Movido a Eliminados)"
                    modelo = model_name.title()
                    user = request.user
                    user_group = user.groups.first().name if user.groups.exists() else "Sin grupo asignado"
                    mensaje = f"""
                    El usuario {user.get_full_name()} (Grupo: {user_group}) ha movido el siguiente registro a la tabla de Eliminados:

                    Acción: {accion}
                    Modelo: {modelo}
                    Datos del Registro:
                    - Activo: {instance.activo}
                    - Marca: {instance.marca}
                    - Modelo: {instance.modelo}
                    - N° Serie: {instance.n_serie}
                    - UNIVE: {instance.unive}
                    - BDO: {instance.bdo}
                    - Estado: {instance.estado}
                    - Creado por: {instance.creado_por.get_full_name() if instance.creado_por else 'N/A'}
                    - Fecha Creación: {instance.fecha_creacion}
                    - Fecha Última Modificación: {instance.fecha_modificacion}
                    - NetBIOS: {getattr(instance, 'netbios', '')}
                    - Ubicación: {getattr(instance, 'ubicacion', '')}
"""
                    if model_name == 'monitor':
                        mensaje += f"""\
                    - Asignado a: {getattr(instance, 'asignado_a', 'N/A')}
"""

                    print("DEBUG: Intentando enviar correo de notificación de eliminación.") # Debug print
                    try:
                        print("DEBUG: Llamando a enviar_notificacion_asunto.") # Debug print
                        enviar_notificacion_asunto(
                            asunto=f"Registro Movido a Eliminados - {modelo}",
                            mensaje=mensaje,
                            destinatarios=settings.EMAIL_RECIPIENTS
                        )
                        print("DEBUG: Llamada a enviar_notificacion_asunto completada.") # Debug print
                        messages.success(request, f'{model_name.title()} movido correctamente a la tabla de Eliminados y correo enviado.')
                    except Exception as e:
                        print(f"DEBUG: Error capturado al enviar correo: {str(e)}") # Debug print
                        logger.error(f'Error al enviar el correo de notificación: {str(e)}')
                        messages.error(request, f'Error al enviar el correo de notificación: {str(e)}')

                else:
                    raise IntegrityError("Error al guardar en la tabla de Eliminados")

        except IntegrityError as e:
            logger.error(f'Error de integridad al mover el registro {model_name} a Eliminados: {e}')
            messages.error(request, 'Error al mover el registro: Problema de integridad al guardar en la tabla Eliminados.')
        except Exception as e:
            logger.error(f'Error al mover el registro {model_name} a Eliminados: {e}')
            messages.error(request, f'Error al mover el registro: {str(e)}')

        return redirect(self.success_url)




# -------- VISTAS DE ALL IN ONE ADMINISTRATIVOS --------

@add_group_name_to_context
class AllInOneAdminView(LoginRequiredMixin, ListView):
    """Vista para listar All In One Administrativos"""
    model = AllInOneAdmins
    template_name = 'modulos/all_in_one_adm.html'
    context_object_name = 'all_in_ones_admins'
    paginate_by = 25
    ordering = ['ubicacion', '-fecha_creacion']

    def get_queryset(self):
        """Obtiene y filtra la lista según búsqueda"""
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        search_by_pk = self.request.GET.get('search_by_pk', 'false').lower() == 'true'

        # Validar si el parámetro es búsqueda por PK y si el input es numérico
        if search_by_pk:
            if search_query.isdigit():  # Solo buscar por PK si es un número
                return queryset.filter(pk=int(search_query))
            else:
                # Si no es numérico, devolver queryset vacío
                return queryset.none()
       
        if search_query:
            queryset = queryset.filter(
                Q(estado__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |     
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(ubicacion__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query) |
                Q(fecha_creacion__icontains=search_query)
            )
       
        return queryset.select_related('creado_por').order_by('ubicacion', '-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '').strip()
        return context

# -------- CONTINUACIÓN DE VISTAS ALL IN ONE ADMINISTRATIVOS --------

@add_group_name_to_context
class Add_AllInOneAdminView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo All In One Administrativo"""
    model = AllInOneAdmins
    template_name = 'modulos/add_all_in_one_adm.html'
    form_class = AllInOneAdminsForm
    success_url = reverse_lazy('all_in_one_adm')

    def test_func(self):
        """Verifica permisos: solo ADR y Operadores pueden agregar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega lista de All In One Administrativos al contexto"""
        context = super().get_context_data(**kwargs)
        context['all_in_ones_adms'] = AllInOneAdmins.objects.all()
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con validaciones adicionales y envía notificación"""
        try:
            if form.is_valid():
                # Validación del número de serie
                n_serie = form.cleaned_data.get('n_serie')
                if not n_serie:
                    form.add_error('n_serie', 'El número de serie es requerido')
                    return self.form_invalid(form)
                if AllInOneAdmins.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este Número de Serie ya existe')
                    return self.form_invalid(form)

                # Validación de campos obligatorios
                required_fields = {
                    'modelo': 'El modelo es requerido',
                    'unive': 'El UNIVE es requerido',
                    'netbios': 'El NetBios es requerido',
                    # 'bdo': 'El BDO es requerido'  # Eliminamos BDO de los campos requeridos en la vista
                }
                for field, error_message in required_fields.items():
                    if not form.cleaned_data.get(field):
                        form.add_error(field, error_message)
                        return self.form_invalid(form)

                # Validación del formato BDO (exactamente igual que en AllInOne)
                bdo = form.cleaned_data.get('bdo')
                try:
                    int(bdo)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO solo debe contener números.')
                    return self.form_invalid(form)

                # Guarda con el usuario actual
                form.instance.creado_por = self.request.user
                instance = form.save()  # Aquí se guarda la instancia para ser utilizada en el correo electrónico

                # Enviar notificación por correo
                accion = "Registro Agregado"  # Cambiar esto según la operación (p. ej., "Registro Eliminado" o "Registro Modificado")
                modelo = "All In One Administrativo"  # Cambiar esto según el modelo en cuestión
                mensaje = f"""
                El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
                Acción: {accion}
                Modelo: {modelo}
                Datos:
                - Estado: {instance.estado}
                - Activo: {modelo}
                - Marca: {instance.marca}
                - Modelo: {instance.modelo}
                - N° Serie: {instance.n_serie}
                - UNIVE: {instance.unive}
                - NetBIOS: {instance.netbios}
                - BDO: {instance.bdo}
                - Ubicación: {instance.ubicacion}
                """

                enviar_notificacion_asunto(
                    asunto="Registro Agregado en All In One Administrativo",  # Cambia el asunto según la acción
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )

                messages.success(self.request, 'All In One Administrativo se ha guardado correctamente.')
                return super().form_valid(form)

            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)

    

@add_group_name_to_context
class Edit_AllInOneAdmView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar All In One Administrativo existente"""
    model = AllInOneAdmins
    template_name = 'modulos/edit_all_in_one_adm.html'
    form_class = AllInOneAdminsForm
    success_url = reverse_lazy('all_in_one_adm')

    def test_func(self):
        """Solo ADR y Operadores puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['opciones_marca_all_in_one'] = opciones_marca_all_in_one
        # Añadir el nombre del modelo para el botón de eliminar
        context['model_name'] = self.model._meta.model_name
        return context

    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación"""
        # Obtener los datos originales del objeto antes de la modificación
        original_instance = AllInOneAdmins.objects.get(pk=form.instance.pk)
        original_data = model_to_dict(original_instance)

        # Desconectar temporalmente la señal para evitar duplicados en el historial
        from django.db.models.signals import pre_save
        from adr.signals import registrar_cambios
        pre_save.disconnect(registrar_cambios, sender=AllInOneAdmins)

        try:
            # Validación de número de serie único
            n_serie = form.cleaned_data.get('n_serie')
            if AllInOneAdmins.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'modelo': 'El modelo es requerido',
                'unive': 'El UNIVE es requerido',
                'netbios': 'El NetBios es requerido',
                # 'bdo': 'El BDO es requerido'  # Añadimos BDO como campo requerido
            }
            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación del formato BDO
            bdo = form.cleaned_data.get('bdo')
            if bdo: # Solo validar si BDO no está vacío
                try:
                    int(bdo)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO solo debe contener números')
                    return self.form_invalid(form)

            # Mantiene el creador original
            original_creado_por = form.instance.creado_por
            instance = form.save(commit=False)
            instance.creado_por = original_creado_por or self.request.user

            # Recopilar cambios para registrar en HistorialCambios
            cambios = {}
            for field_name, new_value in form.cleaned_data.items():
                # Excluir campos que no necesitan ser registrados o que se manejan automáticamente
                if field_name in ['creado_por', 'fecha_creacion', 'id']:
                    continue

                old_value = original_data.get(field_name)

                # Convertir tipos complejos a string para comparación y almacenamiento
                if isinstance(old_value, (Decimal, UUID, datetime)):
                    old_value = str(old_value)
                if isinstance(new_value, (Decimal, UUID, datetime)):
                    new_value = str(new_value)

                # Manejar casos donde los valores pueden ser None o cadenas vacías
                old_value_str = str(old_value) if old_value is not None else ''
                new_value_str = str(new_value) if new_value is not None else ''

                if old_value_str.strip().lower() != new_value_str.strip().lower():
                    cambios[field_name] = {
                        'anterior': old_value_str,
                        'nuevo': new_value_str
                    }

            # Crear registros individuales en HistorialCambios si hay cambios
            if cambios:
                for field_name, vals in cambios.items():
                    HistorialCambios.objects.create(
                        modelo=self.model.__name__,
                        objeto_id=instance.pk,
                        usuario=self.request.user,
                        campo_modificado=field_name.replace('_', ' ').capitalize(), # Usar el nombre del campo como campo_modificado
                        valor_anterior=vals['anterior'],
                        valor_nuevo=vals['nuevo']
                    )

            # Preparar mensaje para notificación por correo
            accion = "Registro Modificado"
            modelo_nombre = "All In One Administrativo"
            mensaje_cuerpo = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo_nombre}
            ID Objeto: {instance.pk}
            """
            if cambios:
                mensaje_cuerpo += "\nDatos Modificados:\n"
                for field_name, vals in cambios.items():
                    mensaje_cuerpo += f"- {field_name.replace('_', ' ').capitalize()}: De '{vals['anterior']}' a '{vals['nuevo']}'\n"
            else:
                 mensaje_cuerpo += "\nNo se detectaron cambios significativos en los campos monitoreados."


            # Enviar notificación por correo
            asunto_correo = f"Registro Modificado en {modelo_nombre}"
            mensaje = mensaje_cuerpo

            try:
                enviar_notificacion_asunto(
                    asunto=asunto_correo,
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )
            except Exception as e:
                messages.error(self.request, f'Error al enviar el correo de notificación: {str(e)}')
                # No retornar self.form_invalid(form) aquí, ya que el formulario es válido.
                # El error es solo en el envío del correo.

            messages.success(self.request, 'Registro All In One Administradores se ha modificado correctamente.')
            return super().form_valid(form)

        finally:
            # Reconectar la señal
            pre_save.connect(registrar_cambios, sender=AllInOneAdmins)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        print("Errores del formulario:", form.errors) # Añadido para depuración
        messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
        return super().form_invalid(form)

        # Validación de número de serie único
        n_serie = form.cleaned_data.get('n_serie')
        if AllInOneAdmins.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        # Validación de campos obligatorios
        required_fields = {
            'modelo': 'El modelo es requerido',
            'unive': 'El UNIVE es requerido',
            'netbios': 'El NetBios es requerido',
            # 'bdo': 'El BDO es requerido'  # Añadimos BDO como campo requerido
        }
        for field, error_message in required_fields.items():
            if not form.cleaned_data.get(field):
                form.add_error(field, error_message)
                return self.form_invalid(form)

        # Validación del formato BDO
        bdo = form.cleaned_data.get('bdo')
        if bdo: # Solo validar si BDO no está vacío
            try:
                int(bdo)
            except ValueError:
                form.add_error('bdo', 'El campo BDO solo debe contener números')
                return self.form_invalid(form)

        # Mantiene el creador original
        original_creado_por = form.instance.creado_por
        instance = form.save(commit=False)
        instance.creado_por = original_creado_por or self.request.user

        # Recopilar cambios para registrar en HistorialCambios
        cambios = {}
        for field_name, new_value in form.cleaned_data.items():
            # Excluir campos que no necesitan ser registrados o que se manejan automáticamente
            if field_name in ['creado_por', 'fecha_creacion', 'id']:
                continue

            old_value = original_data.get(field_name)

            # Convertir tipos complejos a string para comparación y almacenamiento
            if isinstance(old_value, (Decimal, UUID, datetime)):
                old_value = str(old_value)
            if isinstance(new_value, (Decimal, UUID, datetime)):
                new_value = str(new_value)

            # Manejar casos donde los valores pueden ser None o cadenas vacías
            old_value_str = str(old_value) if old_value is not None else ''
            new_value_str = str(new_value) if new_value is not None else ''

            if old_value_str != new_value_str:
                cambios[field_name] = {
                    'anterior': old_value_str,
                    'nuevo': new_value_str
                }

        # Crear un único registro en HistorialCambios si hay cambios
        if cambios:
            campo_modificado_summary = f"Múltiples campos modificados ({', '.join(cambios.keys())})"
            valor_anterior_detail = "\n".join([f"{field}: {vals['anterior']}" for field, vals in cambios.items()])
            valor_nuevo_detail = "\n".join([f"{field}: {vals['nuevo']}" for field, vals in cambios.items()])

            HistorialCambios.objects.create(
                modelo=self.model.__name__,
                objeto_id=instance.pk,
                usuario=self.request.user,
                campo_modificado=campo_modificado_summary,
                valor_anterior=valor_anterior_detail,
                valor_nuevo=valor_nuevo_detail
            )

        # Preparar mensaje para notificación por correo
        accion = "Registro Modificado"
        modelo_nombre = "All In One Administrativo"
        mensaje_cuerpo = f"""
        El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
        Acción: {accion}
        Modelo: {modelo_nombre}
        ID Objeto: {instance.pk}
        """
        if cambios:
            mensaje_cuerpo += "\nDatos Modificados:\n"
            for field_name, vals in cambios.items():
                mensaje_cuerpo += f"- {field_name.replace('_', ' ').capitalize()}: De '{vals['anterior']}' a '{vals['nuevo']}'\n"
        else:
             mensaje_cuerpo += "\nNo se detectaron cambios significativos en los campos monitoreados."


        # Enviar notificación por correo
        asunto_correo = f"Registro Modificado en {modelo_nombre}"
        mensaje = mensaje_cuerpo

        try:
            enviar_notificacion_asunto(
                asunto=asunto_correo,
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )
        except Exception as e:
            messages.error(self.request, f'Error al enviar el correo de notificación: {str(e)}')
            # No retornar self.form_invalid(form) aquí, ya que el formulario es válido.
            # El error es solo en el envío del correo.

        messages.success(self.request, 'Registro All In One Administradores se ha modificado correctamente.')
        return super().form_valid(form)

        try:
            enviar_notificacion_asunto(
                asunto="Registro Modificado en All In One Administrativo",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )
        except Exception as e:
            messages.error(self.request, f'Error al enviar el correo de notificación: {str(e)}')
            # No retornar self.form_invalid(form) aquí, ya que el formulario es válido.
            # El error es solo en el envío del correo.

        messages.success(self.request, 'Registro All In One Administradores se ha modificado correctamente.')
        return super().form_valid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        print("Errores del formulario:", form.errors) # Añadido para depuración
        messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
        return super().form_invalid(form)




# -------- VISTAS DE NOTEBOOKS --------

@add_group_name_to_context
class NotebooksView(LoginRequiredMixin, ListView):
    """Vista para listar todos los Notebooks"""
    model = Notebook
    template_name = 'modulos/notebooks.html'
    context_object_name = 'notebooks'
    paginate_by = 25
    ordering = ['ubicacion', '-fecha_creacion']

    def get_queryset(self):
        """Obtiene y filtra la lista según búsqueda"""
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        search_by_pk = self.request.GET.get('search_by_pk', 'false').lower() == 'true'

        # Validar si el parámetro es búsqueda por PK y si el input es numérico
        if search_by_pk:
            if search_query.isdigit():  # Solo buscar por PK si es un número
                return queryset.filter(pk=int(search_query))
            else:
                # Si no es numérico, devolver queryset vacío
                return queryset.none()
        
        if search_query:
            queryset = queryset.filter(
                Q(estado__icontains=search_query) |
                Q(asignado_a__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(ubicacion__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query) |
                Q(fecha_creacion__icontains=search_query)
            )
        
        return queryset.select_related('creado_por').order_by('ubicacion', '-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '').strip()
        return context

# -------- CONTINUACIÓN DE VISTAS DE NOTEBOOKS --------

@add_group_name_to_context
class AddNotebooksView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo Notebook"""
    model = Notebook
    template_name = 'modulos/add_notebooks.html'
    form_class = NotebooksForm
    success_url = reverse_lazy('notebooks')

    def test_func(self):
        """Verifica permisos: solo ADR, Operadores y Auxiliares pueden agregar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega lista de Notebooks al contexto"""
        context = super().get_context_data(**kwargs)
        context['notebooks'] = Notebook.objects.all()
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con validaciones adicionales y envía notificación"""
        try:
            if form.is_valid():
                # Validación del número de serie
                n_serie = form.cleaned_data.get('n_serie')
                if not n_serie:
                    form.add_error('n_serie', 'El número de serie es requerido')
                    return self.form_invalid(form)
                if Notebook.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este Número de Serie ya existe')
                    return self.form_invalid(form)

                # Validación de campos obligatorios
                required_fields = {
                    'modelo': 'El modelo es requerido',
                    'unive': 'El UNIVE es requerido',
                    'netbios': 'El NetBios es requerido',
                    # Eliminada la validación explícita para BDO
                }
                for field, error_message in required_fields.items():
                    if not form.cleaned_data.get(field):
                        form.add_error(field, error_message)
                        return self.form_invalid(form)

                # Validación del formato BDO (mantener esta validación si se ingresa un valor)
                bdo = form.cleaned_data.get('bdo')
                if bdo is not None and bdo != '': # Solo validar si se ingresó un valor
                    try:
                        int(bdo)
                    except ValueError:
                        form.add_error('bdo', 'El campo BDO solo debe contener números.')
                        return self.form_invalid(form)

                # Guarda con el usuario actual
                logger.debug(f"User before saving: {self.request.user}")
                form.instance.creado_por = self.request.user
                logger.debug(f"creado_por before saving: {form.instance.creado_por}")
                instance = form.save()
                logger.debug(f"creado_por after saving: {instance.creado_por}")

                # Enviar notificación por correo
                accion = "Registro Agregado"
                modelo = "Notebook"
                mensaje = f"""
                El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
                Acción: {accion}
                Modelo: {modelo}
                Datos:
                - Estado: {instance.estado}
                - Activo: {modelo}
                - Marca: {instance.marca}
                - Modelo: {instance.modelo}
                - N° Serie: {instance.n_serie}
                - UNIVE: {instance.unive}
                - NetBIOS: {instance.netbios}
                - BDO: {instance.bdo}
                - Ubicación: {instance.ubicacion}
                - Asignado a: {instance.asignado_a}
                """

                enviar_notificacion_asunto(
                    asunto="Registro Agregado en Notebook",
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )

                messages.success(self.request, 'Notebook se ha guardado correctamente.')
                return super().form_valid(form)
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)


@add_group_name_to_context
class Edit_NotebooksView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar Notebook existente"""
    model = Notebook
    template_name = 'modulos/edit_notebook.html'
    form_class = NotebooksForm
    success_url = reverse_lazy('notebooks')

    def test_func(self):
        """Solo ADR y Operadores pueden editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega datos adicionales al contexto"""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Modificar Notebook'
        context['opciones_marca_notebook'] = opciones_marca_notebook
        # Añadir el nombre del modelo para el botón de eliminar
        context['model_name'] = self.model._meta.model_name
        return context

    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación por correo"""
        try:
            # Validación de número de serie único
            n_serie = form.cleaned_data.get('n_serie')
            if Notebook.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'modelo': 'El modelo es requerido',
                'unive': 'El UNIVE es requerido',
                'netbios': 'El NetBios es requerido',
                # 'bdo': 'El BDO es requerido'
            }
            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación del formato BDO
            bdo = form.cleaned_data.get('bdo')
            try:
                int(bdo)
            except ValueError:
                form.add_error('bdo', 'El campo BDO solo debe contener números.')
                return self.form_invalid(form)

            # Mantiene el creador original
            original_creado_por = form.instance.creado_por
            instance = form.save(commit=False)
            instance.creado_por = original_creado_por or self.request.user
            instance.save()

            # Enviar notificación por correo
            accion = "Registro Modificado"
            modelo = "Notebook"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos:
            - Estado: {instance.estado}
            - Activo: {modelo}
            - Marca: {instance.marca}
            - Modelo: {instance.modelo}
            - N° Serie: {instance.n_serie}
            - UNIVE: {instance.unive}
            - NetBIOS: {instance.netbios}
            - BDO: {instance.bdo}
            - Ubicación: {instance.ubicacion}
            - Asignado a: {instance.asignado_a}
            """

            enviar_notificacion_asunto(
                asunto="Registro Modificado en Notebook",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            messages.success(self.request, 'Registro Notebook se ha modificado correctamente.')
            return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)




# -------- VISTAS DE MINI PC --------

@add_group_name_to_context
class MiniPCView(LoginRequiredMixin, ListView):
    """Vista para listar todos los Mini PC"""
    model = MiniPC
    template_name = 'modulos/mini_pc.html'
    context_object_name = 'minis_pcs'
    paginate_by = 25
    ordering = ['ubicacion', '-fecha_creacion']

    def get_queryset(self):
        """Obtiene y filtra la lista según búsqueda"""
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        search_by_pk = self.request.GET.get('search_by_pk', 'false').lower() == 'true'

        # Validar si el parámetro es búsqueda por PK y si el input es numérico
        if search_by_pk:
            if search_query.isdigit():  # Solo buscar por PK si es un número
                return queryset.filter(pk=int(search_query))
            else:
                # Si no es numérico, devolver queryset vacío
                return queryset.none()
        
        if search_query:
            queryset = queryset.filter(
                Q(activo__icontains=search_query) |
                Q(estado__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(ubicacion__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query) |
                Q(fecha_creacion__icontains=search_query)
            )
        
        return queryset.select_related('creado_por').order_by('ubicacion', '-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '').strip()
        return context

# -------- CONTINUACIÓN DE VISTAS DE MINI PC --------
@add_group_name_to_context
class AddMiniPCView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo Mini PC"""
    model = MiniPC
    template_name = 'modulos/add_mini_pc.html'
    form_class = MiniPCForm
    success_url = reverse_lazy('mini_pc')

    def test_func(self):
        """Verifica permisos: solo ADR y Operadores pueden agregar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega lista de Mini PC al contexto"""
        context = super().get_context_data(**kwargs)
        context['mini_pc'] = MiniPC.objects.all()
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con validaciones adicionales y envía notificación por correo"""
        try:
            if form.is_valid():
                # Validación del número de serie
                n_serie = form.cleaned_data.get('n_serie')
                if not n_serie:
                    form.add_error('n_serie', 'El número de serie es requerido')
                    return self.form_invalid(form)
                if MiniPC.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este Número de Serie ya existe')
                    return self.form_invalid(form)

                # Validación de campos obligatorios
                required_fields = {
                    'modelo': 'El modelo es requerido',
                    'unive': 'El UNIVE es requerido',
                    'bdo': 'El BDO es requerido'
                }
                for field, error_message in required_fields.items():
                    value = form.cleaned_data.get(field)
                    if value in [None, '']:  # Permite 0 como válido
                        form.add_error(field, error_message)
                        return self.form_invalid(form)

                # Validación del formato BDO
                bdo = form.cleaned_data.get('bdo')
                try:
                    int(bdo)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO solo debe contener números.')
                    return self.form_invalid(form)

                # Guarda con el usuario actual
                form.instance.creado_por = self.request.user
                instance = form.save()

                # Enviar notificación por correo
                accion = "Registro Agregado"
                modelo = "Mini PC"
                mensaje = f"""
                El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
                Acción: {accion}
                Modelo: {modelo}
                Datos:
                - Estado: {instance.estado}
                - Activo: {modelo}
                - Marca: {instance.marca}
                - Modelo: {instance.modelo}
                - N° Serie: {instance.n_serie}
                - UNIVE: {instance.unive}
                - NetBIOS: {instance.netbios}
                - BDO: {instance.bdo}
                - Ubicación: {instance.ubicacion}
                """

                enviar_notificacion_asunto(
                    asunto="Registro Agregado en Mini PC",
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )

                messages.success(self.request, 'Mini PC se ha guardado correctamente.')
                return super().form_valid(form)
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)


@add_group_name_to_context
class Edit_MiniPCView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar Mini PC existente"""
    model = MiniPC
    template_name = 'modulos/edit_mini_pc.html'
    form_class = MiniPCForm
    success_url = reverse_lazy('mini_pc')

    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')
        
    def get_context_data(self, **kwargs):
        """Agrega datos adicionales al contexto"""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Modificar Mini PC'
        context['opciones_marca_mini_pc'] = opciones_marca_mini_pc
        # Añadir el nombre del modelo para el botón de eliminar
        context['model_name'] = self.model._meta.model_name
        return context

    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación por correo"""
        try:
            # Validación de número de serie único
            n_serie = form.cleaned_data.get('n_serie')
            if MiniPC.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'modelo': 'El modelo es requerido',
                'unive': 'El UNIVE es requerido',
                # 'bdo': 'El BDO es requerido'
            }
            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación del formato BDO
            # Validación del formato BDO
            bdo = form.cleaned_data.get('bdo')
            if bdo: # Add check if bdo is not None or empty
                try:
                    int(bdo)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO solo debe contener números.')
                    return self.form_invalid(form)

            # Mantiene el creador original
            original_creado_por = form.instance.creado_por
            instance = form.save(commit=False)
            instance.creado_por = original_creado_por or self.request.user
            instance.save()

            # Enviar notificación por correo
            accion = "Registro Modificado"
            modelo = "Mini PC"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos:
            - Estado: {instance.estado}
            - Activo: {modelo}
            - Marca: {instance.marca}
            - Modelo: {instance.modelo}
            - N° Serie: {instance.n_serie}
            - UNIVE: {instance.unive}
            - NetBIOS: {instance.netbios}
            - BDO: {instance.bdo}
            - Ubicación: {instance.ubicacion}
            """

            enviar_notificacion_asunto(
                asunto="Registro Modificado en Mini PC",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            messages.success(self.request, 'Registro Mini PC se ha modificado correctamente.')
            return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)



# -------- VISTAS DE PROYECTORES --------

@add_group_name_to_context
class ProyectoresView(LoginRequiredMixin, ListView):
    """Vista para listar todos los Proyectores"""
    model = Proyectores
    template_name = 'modulos/proyectores.html'
    context_object_name = 'proyectores'
    paginate_by = 25
    ordering = ['ubicacion', '-fecha_creacion']

    def get_queryset(self):
        """Obtiene y filtra la lista según búsqueda"""
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        search_by_pk = self.request.GET.get('search_by_pk', 'false').lower() == 'true'

        # Validar si el parámetro es búsqueda por PK y si el input es numérico
        if search_by_pk:
            if search_query.isdigit():  # Solo buscar por PK si es un número
                return queryset.filter(pk=int(search_query))
            else:
                # Si no es numérico, devolver queryset vacío
                return queryset.none()
       
        if search_query:
            queryset = queryset.filter(                
                Q(estado__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(ubicacion__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query) |
                Q(fecha_creacion__icontains=search_query)
            )
       
        return queryset.select_related('creado_por').order_by('ubicacion', '-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '').strip()
        return context

# -------- CONTINUACIÓN DE VISTAS DE PROY@add_group_name_to_context
@add_group_name_to_context
class AddProyectorView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo Proyector"""
    model = Proyectores
    template_name = 'modulos/add_proyector.html'
    form_class = ProyectoresForm
    success_url = reverse_lazy('proyectores')

    def test_func(self):
        """Verifica permisos: solo ADR y Operadores pueden agregar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega lista de Proyectores al contexto"""
        context = super().get_context_data(**kwargs)
        context['proyectores'] = Proyectores.objects.all()
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con validaciones adicionales y envía notificación por correo"""
        try:
            if form.is_valid():
                # Validación del número de serie
                n_serie = form.cleaned_data.get('n_serie')
                if not n_serie:
                    form.add_error('n_serie', 'El número de serie es requerido')
                    return self.form_invalid(form)
                if Proyectores.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este Número de Serie ya existe')
                    return self.form_invalid(form)

                # Validación de campos obligatorios
                required_fields = {
                    'modelo': 'El modelo es requerido',
                }
                for field, error_message in required_fields.items():
                    if not form.cleaned_data.get(field):
                        form.add_error(field, error_message)
                        return self.form_invalid(form)

                # Guarda con el usuario actual
                form.instance.creado_por = self.request.user
                instance = form.save()  # Guardamos la instancia del proyector

                # Enviar notificación por correo
                accion = "Registro Agregado"
                modelo = "Proyector"
                mensaje = f"""
                El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
                Acción: {accion}
                Modelo: {modelo}
                Datos:
                - Estado: {instance.estado}
                - Marca: {instance.marca}
                - Modelo: {instance.modelo}
                - N° Serie: {instance.n_serie}
                - Ubicación: {instance.ubicacion}
                """

                enviar_notificacion_asunto(
                    asunto="Registro Agregado en Proyector",
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS
                )

                messages.success(self.request, 'Proyector se ha guardado correctamente y la notificación se ha enviado.')
                return super().form_valid(form)

            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'Ha ocurrido un error inesperado: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Muestra los errores específicos del formulario"""
        error_messages = []
        for field, errors in form.errors.items():
            for error in errors:
                error_messages.append(f"Error en el campo '{field}': {error}")

        # Mostrar todos los errores encontrados
        if error_messages:
            detailed_errors = "\n".join(error_messages)
            messages.error(self.request, f'Errores encontrados en el formulario:\n{detailed_errors}')
        else:
            messages.error(self.request, 'Ha ocurrido un error desconocido al procesar el formulario.')

        return super().form_invalid(form)





    
@add_group_name_to_context
class Edit_ProyectorView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar Proyector existente"""
    model = Proyectores
    template_name = 'modulos/edit_proyector.html'
    form_class = ProyectoresForm
    success_url = reverse_lazy('proyectores')

    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega datos adicionales al contexto"""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Modificar Proyector'
        context['opciones_marca_proyector'] = opciones_marca_proyector
        # Añadir el nombre del modelo para el botón de eliminar
        context['model_name'] = self.model._meta.model_name
        return context

    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación por correo"""
        try:
            # Validación del número de serie único
            n_serie = form.cleaned_data.get('n_serie')
            if Proyectores.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'estado': 'El estado es requerido',
                'marca': 'La marca es requerida',
                'modelo': 'El modelo es requerido',
                'ubicacion': 'El edificio es requerido'
            }
            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Verificación de marca existente
            marca = form.cleaned_data.get('marca')
            if marca == '':
                form.add_error('marca', 'Debe seleccionar una marca')
                return self.form_invalid(form)

            # Verificación de ubicación existente
            ubicacion = form.cleaned_data.get('ubicacion')
            if ubicacion == '':
                form.add_error('ubicacion', 'Debe seleccionar un edificio')
                return self.form_invalid(form)

            # Mantiene el creador original
            original_creado_por = form.instance.creado_por
            instance = form.save(commit=False)
            instance.creado_por = original_creado_por or self.request.user
            instance.save()

            # Enviar notificación por correo
            accion = "Registro Modificado"
            modelo = "Proyector"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos:
            - Estado: {instance.estado}
            - Marca: {instance.marca}
            - Modelo: {instance.modelo}
            - N° Serie: {instance.n_serie}
            - Ubicación: {instance.ubicacion}
            """

            enviar_notificacion_asunto(
                asunto="Registro Modificado en Proyector",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            messages.success(self.request, 'Registro Proyector se ha modificado correctamente.')
            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)

    


# -------- VISTAS DE BODEGA ADR --------

@add_group_name_to_context
class BodegaADRView(LoginRequiredMixin, ListView):
    """Vista para listar todos los equipos en Bodega ADR"""
    model = BodegaADR
    template_name = 'modulos/bodega_adr.html'
    context_object_name = 'bodegas_adr'
    paginate_by = 25
    ordering = ['activo', '-fecha_creacion']


    def get_queryset(self):
        """Obtiene y filtra la lista según búsqueda"""
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        search_by_pk = self.request.GET.get('search_by_pk', 'false').lower() == 'true'

        # Manejo de búsqueda por pk
        if search_by_pk:
            if search_query.isdigit():  # Validar si es un número
                return queryset.filter(pk=int(search_query))
            else:
                # Si no es un número, no se realiza búsqueda por pk
                return queryset.none()
       
        if search_query:
            queryset = queryset.filter(
                
                Q(ubicacion__icontains=search_query) |
                Q(activo__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query) |
                Q(fecha_creacion__icontains=search_query)
            )
       
        return queryset.select_related('creado_por').order_by('activo', '-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '').strip()
        return context

# -------- CONTINUACIÓN DE VISTAS DE BODEGA ADR --------

@add_group_name_to_context
class AddBodegaADRView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo equipo a Bodega ADR"""
    model = BodegaADR
    template_name = 'modulos/add_bodega_adr.html'
    form_class = BodegaADRForm
    success_url = reverse_lazy('bodega_adr')

    def test_func(self):
        """Verifica permisos para agregar a Bodega"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']

    def handle_no_permission(self):
        return redirect('error')

    def form_valid(self, form):
        """Procesa el formulario con validaciones"""
        try:
            # Validación del número de serie
            n_serie = form.cleaned_data.get('n_serie')
            if not n_serie:
                form.add_error('n_serie', 'El número de serie es requerido')
                return self.form_invalid(form)

            if BodegaADR.objects.filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'activo': 'El tipo de activo es requerido',
                'marca': 'La marca es requerida',
                'modelo': 'El modelo es requerido',
                'ubicacion': 'La ubicación es requerida', # Cambiado de estado_activo a ubicacion
            }

            for field, error_message in required_fields.items():
                value = form.cleaned_data.get(field)
                # Para campos de texto, verificamos si están vacíos después de quitar espacios
                if isinstance(value, str) and not value.strip():
                     form.add_error(field, error_message)
                     return self.form_invalid(form)
                # Para otros tipos de campos obligatorios (aunque en este caso solo son texto)
                elif value is None:
                     form.add_error(field, error_message)
                     return self.form_invalid(form)


            # Validación del formato BDO (ahora opcional)
            bdo = form.cleaned_data.get('bdo')
            if bdo is not None and bdo != '': # Permitir BDO vacío
                try:
                    bdo = int(bdo)
                    if bdo < 0:
                        form.add_error('bdo', 'El BDO debe ser un número positivo o cero.')
                        return self.form_invalid(form)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO solo debe contener números.')
                    return self.form_invalid(form)


            # Guardar el formulario
            form.instance.creado_por = self.request.user
            instance = form.save()

            # Mensaje de éxito si todo va bien
            messages.success(self.request, 'Activo agregado correctamente a Bodega ADR.')

            # Enviar notificación por correo
            accion = "Registro Agregado"
            modelo = "Bodega ADR"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos:
            - Ubicación: {instance.ubicacion} # Cambiado de estado_activo a ubicacion
            - Activo: {instance.activo}
            - Marca: {instance.marca}
            - Modelo: {instance.modelo}
            - N° Serie: {instance.n_serie}
            - UNIVE: {instance.unive}
            - BDO: {instance.bdo if instance.bdo is not None else 'N/A'} # Manejar BDO opcional
            - NetBIOS: {instance.netbios if instance.netbios is not None else 'N/A'} # Manejar NetBIOS opcional
            """

            enviar_notificacion_asunto(
                asunto="Registro Agregado en Bodega ADR",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, f'Ha ocurrido un error inesperado: {str(e)}')
            # No retornar form_invalid aquí para evitar duplicar mensajes de error si ya hay errores de formulario
            return redirect(self.request.path) # Redirigir a la misma página para mostrar el error

    def form_invalid(self, form):
        """Muestra los errores específicos del formulario"""
        error_messages = []
        for field, errors in form.errors.items():
            for error in errors:
                error_messages.append(f"Error en el campo '{field}': {error}")

        # Mostrar todos los errores encontrados
        if error_messages:
            detailed_errors = "\n".join(error_messages)
            messages.error(self.request, f'Errores encontrados en el formulario:\n{detailed_errors}')
        else:
            messages.error(self.request, 'Ha ocurrido un error desconocido al procesar el formulario.')

        return super().form_invalid(form)



@add_group_name_to_context
class Edit_BodegaADRView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar equipo en Bodega ADR"""
    model = BodegaADR
    template_name = 'modulos/edit_bodega_adr.html'
    form_class = BodegaADRForm
    success_url = reverse_lazy('bodega_adr')

    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        return redirect('error')

    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación por correo"""
        try:
            # Validación del número de serie
            n_serie = form.cleaned_data.get('n_serie')
            if not n_serie:
                form.add_error('n_serie', 'El número de serie es requerido')
                return self.form_invalid(form)

            if BodegaADR.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'activo': 'El tipo de activo es requerido',
                'marca': 'La marca es requerida',
                'modelo': 'El modelo es requerido',
                'ubicacion': 'La ubicación es requerida'
            }

            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación de selección en dropdowns
            dropdowns = {
                'activo': 'Debe seleccionar un tipo de activo',
                # 'marca' ya no es un dropdown, se valida en required_fields y en el form.clean_marca
                'ubicacion': 'Debe seleccionar o ingresar una ubicación' # Mensaje ajustado para 'ubicacion'
            }

            # Ajuste de la validación para 'ubicacion' ya que 'marca' no es dropdown
            # y 'estado_activo' ahora es 'ubicacion'
            if not form.cleaned_data.get('activo') or form.cleaned_data.get('activo') == '':
                 form.add_error('activo', dropdowns['activo'])
                 return self.form_invalid(form)
            
            # La validación para 'ubicacion' (antes 'estado_activo') se mantiene si se espera que no esté vacía,
            # lo cual ya está cubierto por required_fields. Si se quisiera validar que sea una opción de un select (lo cual ya no es),
            # esta lógica sería diferente. Como ahora es un TextInput, la validación de 'no vacío' es suficiente aquí.
            # La validación de 'if value == ""' para 'ubicacion' se puede mantener si se quiere asegurar que no sea una cadena vacía
            # incluso si el campo es opcional (aunque 'required_fields' ya lo cubre si es obligatorio).
            # Por simplicidad y dado que 'ubicacion' está en required_fields, esta sección de dropdowns
            # podría simplificarse o eliminarse si 'required_fields' y la limpieza del formulario son suficientes.
            # Mantendremos la validación para 'activo' como ejemplo de dropdown.
            # Si 'ubicacion' (antes 'estado_activo') fuera opcional y se quisiera validar que si se provee no sea vacío,
            # se haría aquí. Pero como es requerido, 'required_fields' lo cubre.
            # Para 'marca', la validación de no vacío está en 'required_fields'.

            # Eliminamos la iteración original de dropdowns y validamos 'activo' explícitamente.
            # La validación de 'ubicacion' como campo requerido ya está en 'required_fields'.
            # La validación de 'marca' como campo requerido ya está en 'required_fields'.

            # Validación del BDO si está presente
            bdo = form.cleaned_data.get('bdo')
            if bdo:
                try:
                    int(bdo)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO debe contener solo números')
                    return self.form_invalid(form)

            # Mantiene el creador original
            original_creado_por = form.instance.creado_por
            instance = form.save(commit=False)
            instance.creado_por = original_creado_por or self.request.user
            instance.save()

            # Enviar notificación por correo
            accion = "Registro Modificado"
            modelo = "Bodega ADR"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos:
            - Ubicación: {instance.ubicacion}
            - Activo: {instance.activo}
            - Marca: {instance.marca}
            - Modelo: {instance.modelo}
            - N° Serie: {instance.n_serie}
            - UNIVE: {instance.unive}
            - BDO: {instance.bdo}
            - NetBIOS: {instance.netbios}
            """

            enviar_notificacion_asunto(
                asunto="Registro Modificado en Bodega ADR",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            # Mensaje de éxito si todo va bien
            messages.success(self.request, 'Registro de Bodega ADR modificado correctamente.')
            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)



# -------- VISTAS DE AZOTEA --------

@add_group_name_to_context
class AzoteaView(LoginRequiredMixin, ListView):
    """Vista para listar equipos en Azotea"""
    model = Azotea
    template_name = 'modulos/azotea_adr.html'
    context_object_name = 'azoteas_adr'
    paginate_by = 25
    ordering = ['activo', '-fecha_creacion']
    

    def get_queryset(self):
        """Obtiene y filtra la lista según búsqueda"""
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '').strip()
        search_by_pk = self.request.GET.get('search_by_pk', 'false').lower() == 'true'

        # Validar si el parámetro es búsqueda por PK y si el input es numérico
        if search_by_pk:
            if search_query.isdigit():  # Solo buscar por PK si es un número
                return queryset.filter(pk=int(search_query))
            else:
                # Si no es numérico, devolver queryset vacío
                return queryset.none()
       
        if search_query:
            queryset = queryset.filter(
                Q(ubicacion__icontains=search_query) |
                Q(activo__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(n_serie__icontains=search_query) |
                Q(unive__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(creado_por__first_name__icontains=search_query) |
                Q(creado_por__last_name__icontains=search_query) |
                Q(fecha_creacion__icontains=search_query)
            )
       
        return queryset.select_related('creado_por').order_by('activo', '-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '').strip()
        return context

# -------- CONTINUACIÓN DE VISTAS DE AZOTEA --------
@add_group_name_to_context
class AddAzoteaView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo equipo a Azotea"""
    model = Azotea
    template_name = 'modulos/add_azotea.html'
    form_class = AzoteaForm
    success_url = reverse_lazy('azotea_adr')

    def test_func(self):
        """Verifica permisos: solo ADR y Operadores pueden agregar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR', 'Auxiliares Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación por correo"""
        try:
            # Validación del número de serie
            n_serie = form.cleaned_data.get('n_serie')
            if not n_serie:
                form.add_error('n_serie', 'El número de serie es requerido')
                return self.form_invalid(form)

            if Azotea.objects.filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'activo': 'El tipo de activo es requerido',
                'marca': 'La marca es requerida',
                'modelo': 'El modelo es requerido',
                'ubicacion': 'La ubicación es requerida'
            }

            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación de selección en dropdowns
            dropdowns = {
                'activo': 'Debe seleccionar un tipo de activo',
                'marca': 'Debe seleccionar una marca',
                'ubicacion': 'Debe ingresar una ubicación'
            }

            for field, error_message in dropdowns.items():
                value = form.cleaned_data.get(field)
                if value == '':
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación del BDO si está presente
            bdo = form.cleaned_data.get('bdo')
            if bdo:
                try:
                    int(bdo)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO solo debe contener números')
                    return self.form_invalid(form)

            # Guardar con el usuario actual
            form.instance.creado_por = self.request.user
            instance = form.save()

            # Enviar notificación por correo
            accion = "Registro Agregado"
            modelo = "Azotea ADR"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos:
            - Ubicación: {instance.ubicacion}
            - Activo: {instance.activo}
            - Marca: {instance.marca}
            - Modelo: {instance.modelo}
            - N° Serie: {instance.n_serie}
            - UNIVE: {instance.unive}
            - BDO: {instance.bdo}
            """

            enviar_notificacion_asunto(
                asunto="Registro Agregado en Azotea ADR",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            messages.success(self.request, 'Activo agregado correctamente a Azotea ADR.')
            return super().form_valid(form)

        except Exception as e:
            print(f"Error al guardar activo de Azotea: {e}") # Imprimir el error en la consola
            messages.error(self.request, f'Ha ocurrido un error: {e}. Por favor, verifique los datos e intente nuevamente.') # Mostrar el error en la interfaz
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)
    
    
@add_group_name_to_context
class Edit_AzoteaView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar equipo en Azotea"""
    model = Azotea
    template_name = 'modulos/edit_azotea_adr.html'
    form_class = AzoteaForm
    success_url = reverse_lazy('azotea_adr')

    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def get_context_data(self, **kwargs):
        """Agrega datos adicionales al contexto"""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Modificar Registro Azotea'
        return context

    def form_valid(self, form):
        """Procesa el formulario con validaciones y envía notificación por correo"""
        try:
            # Validación del número de serie
            n_serie = form.cleaned_data.get('n_serie')
            if not n_serie:
                form.add_error('n_serie', 'El número de serie es requerido')
                return self.form_invalid(form)

            if Azotea.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
                form.add_error('n_serie', 'Este Número de Serie ya existe')
                return self.form_invalid(form)

            # Validación de campos obligatorios
            required_fields = {
                'activo': 'El tipo de activo es requerido',
                'marca': 'La marca es requerida',
                'modelo': 'El modelo es requerido',
                'ubicacion': 'La ubicación es requerida'
            }

            for field, error_message in required_fields.items():
                if not form.cleaned_data.get(field):
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación de selección en dropdowns
            dropdowns = {
                'activo': 'Debe seleccionar un tipo de activo',
                'marca': 'Debe seleccionar una marca', # Asumiendo que marca en AzoteaForm sigue siendo select
                'ubicacion': 'Debe ingresar una ubicación'
            }

            for field, error_message in dropdowns.items():
                value = form.cleaned_data.get(field)
                if value == '':
                    form.add_error(field, error_message)
                    return self.form_invalid(form)

            # Validación del BDO si está presente
            bdo = form.cleaned_data.get('bdo')
            if bdo:
                try:
                    int(bdo)
                except ValueError:
                    form.add_error('bdo', 'El campo BDO solo debe contener números')
                    return self.form_invalid(form)

            # Mantiene el creador original
            original_creado_por = form.instance.creado_por
            instance = form.save(commit=False)
            instance.creado_por = original_creado_por or self.request.user
            instance.save()

            # Enviar notificación por correo
            accion = "Registro Modificado"
            modelo = "Azotea ADR"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Datos:
            - Ubicación: {instance.ubicacion}
            - Activo: {instance.activo}
            - Marca: {instance.marca}
            - Modelo: {instance.modelo}
            - N° Serie: {instance.n_serie}
            - UNIVE: {instance.unive}
            - BDO: {instance.bdo}
            """

            enviar_notificacion_asunto(
                asunto="Registro Modificado en Azotea ADR",
                mensaje=mensaje,
                destinatarios=settings.EMAIL_RECIPIENTS
            )

            messages.success(self.request, 'Registro de Azotea ADR modificado correctamente.')
            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, 'Ha ocurrido un error. Por favor, verifique los datos e intente nuevamente.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)

# -------- VISTAS DE MONITORES --------

@add_group_name_to_context
class MonitorView(LoginRequiredMixin, ListView):
    """Vista para listar todos los Monitores"""
    model = Monitor
    template_name = 'modulos/monitor.html' # Necesitaremos crear esta plantilla
    context_object_name = 'monitores'
    paginate_by = 15

    def get_queryset(self):
        """
        Sobrescribe get_queryset para cargar explícitamente el usuario relacionado
        y aplicar ordenamiento por defecto.
        """
        queryset = super().get_queryset().select_related('creado_por')
        # Puedes añadir un ordenamiento por defecto si lo deseas, por ejemplo por fecha de creación
        # queryset = queryset.order_by('-fecha_creacion')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name'] = 'monitor'
        context['model_name_plural'] = 'monitores'
        
        # Llamamos a filtrar_y_paginar para obtener los datos de paginación y filtro
        # El queryset base se obtiene de self.object_list que ListView ya ha preparado (y ordenado si self.ordering está definido)
        # o podemos pasar self.model.objects.all() si queremos que filtrar_y_paginar maneje el ordenamiento inicial también.
        # Por consistencia con AzoteaView, dejaremos que filtrar_y_paginar tome self.model.
        
        # Campos para la búsqueda en Monitor (eliminado 'pulgadas')
        search_fields = ['activo', 'marca', 'modelo', 'n_serie', 'unive', 'bdo', 'ubicacion', 'asignado_a', 'creado_por__first_name', 'creado_por__last_name']

        page_obj, filter_ubicacion_actual, ubicaciones_disponibles = filtrar_y_paginar(
            self.request,
            self.model,
            search_fields,
            self.paginate_by
        )
        context['page_obj'] = page_obj
        context['paginator'] = page_obj.paginator
        context['object_list'] = page_obj.object_list # Asegurarse que la lista de objetos en el contexto es la paginada
        context['monitores'] = page_obj.object_list # Actualizar el context_object_name también
        context['filter_ubicacion_actual'] = filter_ubicacion_actual
        context['ubicaciones_disponibles'] = ubicaciones_disponibles
        context['add_url'] = 'add_monitor'
        context['nombre_activo'] = 'Monitor'
        # directamente desde el resultado de la función.
        # El queryset para la paginación se obtiene llamando a get_queryset y luego paginando ese resultado.
        
        # Obtenemos el page_obj y otros datos de filtrar_y_paginar
        page_obj, filter_ubicacion_actual, ubicaciones_disponibles = filtrar_y_paginar(
            self.request,
            self.model,
            ['activo', 'marca', 'modelo', 'n_serie', 'unive', 'bdo', 'ubicacion', 'asignado_a', 'creado_por__first_name', 'creado_por__last_name'],
            self.paginate_by
        )
        context['page_obj'] = page_obj
        context['filter_ubicacion_actual'] = filter_ubicacion_actual
        context['ubicaciones_disponibles'] = ubicaciones_disponibles
        context['add_url'] = 'add_monitor'  # Nombre de la URL para agregar monitor
        context['nombre_activo'] = 'Monitor' # Nombre del activo para el botón
        return context

@add_group_name_to_context
class AddMonitorView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo Monitor"""
    model = Monitor
    form_class = MonitorForm # Necesitaremos crear este formulario
    template_name = 'modulos/add_monitor.html' # Necesitaremos crear esta plantilla
    success_url = reverse_lazy('monitor_list') # Necesitaremos crear esta URL

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name_singular'] = 'Monitor'
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                n_serie = form.cleaned_data.get('n_serie')
                unive = form.cleaned_data.get('unive')

                if Monitor.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este número de serie ya está registrado.')
                    return self.form_invalid(form)

                if unive != "0" and Monitor.objects.filter(unive=unive).exists():
                    form.add_error('unive', 'Este código UNIVE ya está registrado.')
                    return self.form_invalid(form)

                bdo = form.cleaned_data.get('bdo')
                if bdo != 0 and Monitor.objects.filter(bdo=bdo).exists():
                    form.add_error('bdo', 'Este código BDO ya está registrado.')
                    return self.form_invalid(form)

                monitor = form.save(commit=False)
                monitor.creado_por = self.request.user
                monitor.save()

                try:
                    enviar_notificacion_asunto(
                        asunto="Nuevo Monitor Registrado en el Sistema",
                        mensaje=f"""
                        El usuario {self.request.user.get_full_name()} ha agregado un nuevo monitor al sistema.

                        Activo: {monitor.activo}
                        Marca: {monitor.marca}
                        Modelo: {monitor.modelo}
                        N° Serie: {monitor.n_serie}
                        UNIVE: {monitor.unive}
                        BDO: {monitor.bdo}
                        Ubicación: {monitor.ubicacion}
                        """,
                        destinatarios=settings.EMAIL_RECIPIENTS
                    )
                except Exception as e:
                    messages.error(self.request, f'Error al enviar el correo: {str(e)}')

                messages.success(self.request, 'Monitor agregado exitosamente')
                return super().form_valid(form)

        except IntegrityError as e:
            print(f'INTEGRITY ERROR: {e}')
            messages.error(self.request, 'Error de integridad de datos al guardar el monitor.')
            return self.form_invalid(form)

        except Exception as e:
            print(f'ERROR GENERAL: {e}')
            messages.error(self.request, f'Error inesperado al agregar el monitor: {str(e)}')
            return self.form_invalid(form)

        # Esto ya no debería ser necesario, pero lo dejamos por seguridad absoluta
        return self.form_invalid(form)
@add_group_name_to_context
class EditMonitorView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar Monitor existente"""
    model = Monitor
    form_class = MonitorForm
    template_name = 'modulos/edit_monitor.html' # Necesitaremos crear esta plantilla
    success_url = reverse_lazy('monitor_list')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name_singular'] = 'Monitor'
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                monitor = form.save(commit=False)
                # Guardar el estado original para el historial
                original_monitor = Monitor.objects.get(pk=monitor.pk)
                changed_fields = []
                for field in form.changed_data:
                    old_value = getattr(original_monitor, field)
                    new_value = form.cleaned_data[field]
                    if old_value != new_value:
                        HistorialCambios.objects.create(
                            modelo=monitor.__class__.__name__,
                            objeto_id=monitor.pk,
                            usuario=self.request.user,
                            campo_modificado=field,
                            valor_anterior=str(old_value),
                            valor_nuevo=str(new_value)
                        )
                        changed_fields.append(f"{field}: de '{old_value}' a '{new_value}'")
                
                monitor.save()

                # Lógica de notificación por correo si hay cambios
                if changed_fields:
                    accion = "Monitor Editado"
                    nombre_completo_usuario = self.request.user.get_full_name()
                    campos_modificados_str = "\n".join(changed_fields)
                    mensaje = f"""
                    El usuario {nombre_completo_usuario} ha editado un monitor en el sistema.

                    Acción: {accion}
                    Activo: {monitor.activo}
                    Número de Serie: {monitor.n_serie}
                    Cambios realizados:
                    {campos_modificados_str}
                    """
                    try:
                        enviar_notificacion_asunto(
                            asunto="Monitor Editado en el Sistema",
                            mensaje=mensaje,
                            destinatarios=settings.EMAIL_RECIPIENTS
                        )
                    except Exception as e:
                        messages.error(self.request, f'Error al enviar el correo de notificación: {str(e)}')
                
                messages.success(self.request, 'Monitor actualizado exitosamente')
                return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, 'Error de integridad de datos al actualizar el monitor.')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'Error inesperado al actualizar el monitor: {str(e)}')
            return self.form_invalid(form)

@add_group_name_to_context
class MonitorDetailView(LoginRequiredMixin, DetailView):
    model = Monitor
    template_name = 'modulos/detalle_monitor.html' # Necesitaremos crear esta plantilla
    context_object_name = 'monitor'

# -------- VISTAS DE AUDIO --------

@add_group_name_to_context
class AudioView(LoginRequiredMixin, ListView):
    """
    Vista para listar Equipos de Audio, con búsqueda, filtro por estado,
    paginación y botones de editar/eliminar según el grupo del usuario.
    """
    model = Audio
    template_name = 'modulos/audio.html'
    context_object_name = 'page_obj'
    paginate_by = 15

    def get_queryset(self):
        qs = super().get_queryset()

        # 1) Búsqueda global
        q = self.request.GET.get('search', '').strip()
        if q:
            qs = qs.filter(
                Q(activo__icontains=q) |
                Q(marca__icontains=q) |
                Q(modelo__icontains=q) |
                Q(n_serie__icontains=q) |
                Q(unive__icontains=q) |
                Q(bdo__icontains=q) |
                Q(ubicacion__icontains=q) |
                Q(creado_por__first_name__icontains=q) |
                Q(creado_por__last_name__icontains=q)
            )

        # 2) Filtro por estado (case-insensitive)
        estado = self.request.GET.get('estado', '').strip()
        if estado:
            qs = qs.filter(estado__iexact=estado)

        return qs.order_by('n_serie')

    def get_context_data(self, **kwargs):
        # Esto llama al get_context_data del decorator y de ListView
        context = super().get_context_data(**kwargs)

        # Asegurarnos de que “group_name_singular” está en el contexto
        if 'group_name_singular' not in context:
            # Por ejemplo, tomar el primer grupo del usuario
            context['group_name_singular'] = (
                self.request.user.groups.first().name
                if self.request.user.groups.exists()
                else ''
            )

        # Paginación manual: aplicamos get_queryset() ya filtrado
        qs = self.get_queryset()
        paginator = Paginator(qs, self.paginate_by)
        page_number = self.request.GET.get('page')
        context['page_obj']   = paginator.get_page(page_number)
        context['paginator']  = paginator

        # Variables adicionales para la plantilla
        context.update({
            'model_name':          'audio',
            'model_name_plural':   'equipos de audio',
            'add_url':             'add_audio',
            'nombre_activo':       'Equipo de Audio',
        })
        return context
@add_group_name_to_context
class AddAudioView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nuevo Equipo de Audio"""
    model = Audio
    form_class = AudioForm
    template_name = 'modulos/add_audio.html'
    success_url = reverse_lazy('audio_list')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name_singular'] = 'Equipo de Audio'
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                n_serie = form.cleaned_data.get('n_serie')
                unive = form.cleaned_data.get('unive')
                bdo = form.cleaned_data.get('bdo')

                # Validación de duplicados
                if Audio.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este número de serie ya está registrado.')
                    return self.form_invalid(form)

                # Validar UNIVE solo si no es "0" o 0
                if str(unive) not in ["0", 0, "", None] and Audio.objects.filter(unive=unive).exists():
                    form.add_error('unive', 'Este código UNIVE ya existe para un Equipo de Audio.')
                    return self.form_invalid(form)
                if bdo != 0 and Audio.objects.filter(bdo=bdo).exists():
                    form.add_error('bdo', 'Este código BDO ya está registrado.')
                    return self.form_invalid(form)

                # Guardar con usuario
                audio = form.save(commit=False)
                audio.creado_por = self.request.user
                audio.save()

                try:
                    enviar_notificacion_asunto(
                        asunto="Nuevo Equipo de Audio Registrado",
                        mensaje=f"""
                        El usuario {self.request.user.get_full_name()} ha agregado un nuevo equipo de audio al sistema.

                        Activo: {audio.activo}
                        Marca: {audio.marca}
                        Modelo: {audio.modelo}
                        N° Serie: {audio.n_serie}
                        UNIVE: {audio.unive}
                        BDO: {audio.bdo}
                        Ubicación: {audio.ubicacion}
                        """,
                        destinatarios=settings.EMAIL_RECIPIENTS
                    )
                except Exception as e:
                    messages.error(self.request, f'Error al enviar el correo: {str(e)}')

                messages.success(self.request, 'Equipo de Audio agregado exitosamente')
                return super().form_valid(form)

        except IntegrityError as e:
            print(f'INTEGRITY ERROR: {e}')
            messages.error(self.request, 'Error de integridad de datos al guardar el equipo de audio.')
            return self.form_invalid(form)

        except Exception as e:
            print(f'ERROR GENERAL: {e}')
            messages.error(self.request, f'Error inesperado al agregar el equipo de audio: {str(e)}')
            return self.form_invalid(form)

        return self.form_invalid(form)

@add_group_name_to_context
class EditAudioView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar Equipo de Audio existente"""
    model = Audio
    form_class = AudioForm
    template_name = 'modulos/edit_audio.html' # Necesitaremos crear esta plantilla
    success_url = reverse_lazy('audio_list')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name_singular'] = 'Equipo de Audio'
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                audio = form.save(commit=False)
                original_audio = Audio.objects.get(pk=audio.pk)
                changed_fields = []
                for field in form.changed_data:
                    old_value = getattr(original_audio, field)
                    new_value = form.cleaned_data[field]
                    if old_value != new_value:
                        HistorialCambios.objects.create(
                            modelo=audio.__class__.__name__,
                            objeto_id=audio.pk,
                            usuario=self.request.user,
                            campo_modificado=field,
                            valor_anterior=str(old_value),
                            valor_nuevo=str(new_value)
                        )
                        changed_fields.append(f"{field}: de '{old_value}' a '{new_value}'")
                
                audio.save()

                if changed_fields:
                    accion = "Equipo de Audio Editado"
                    nombre_completo_usuario = self.request.user.get_full_name()
                    campos_modificados_str = "\n".join(changed_fields)
                    mensaje = f"""
                    El usuario {nombre_completo_usuario} ha editado un equipo de audio en el sistema.

                    Acción: {accion}
                    Activo: {audio.activo}
                    Número de Serie: {audio.n_serie}
                    Cambios realizados:
                    {campos_modificados_str}
                    """
                    try:
                        enviar_notificacion_asunto(
                            asunto="Equipo de Audio Editado en el Sistema",
                            mensaje=mensaje,
                            destinatarios=settings.EMAIL_RECIPIENTS
                        )
                    except Exception as e:
                        messages.error(self.request, f'Error al enviar el correo de notificación: {str(e)}')
                
                messages.success(self.request, 'Equipo de Audio actualizado exitosamente')
                return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, 'Error de integridad de datos al actualizar el equipo de audio.')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'Error inesperado al actualizar el equipo de audio: {str(e)}')
            return self.form_invalid(form)

@add_group_name_to_context
class AudioDetailView(LoginRequiredMixin, DetailView):
    model = Audio
    template_name = 'modulos/detalle_audio.html' # Necesitaremos crear esta plantilla
    context_object_name = 'audio'

# -------- VISTAS DE TABLETS --------

@add_group_name_to_context
class TabletView(LoginRequiredMixin, ListView):
    model = Tablet
    template_name = 'modulos/tablet.html'
    context_object_name = 'page_obj'   # ahora page_obj, para coincidir con la plantilla
    paginate_by = 15

    def get_queryset(self):
        qs = super().get_queryset()

        # 1) Búsqueda global
        q = self.request.GET.get('search', '').strip()
        if q:
            qs = qs.filter(
                Q(activo__icontains=q)       |
                Q(marca__icontains=q)        |
                Q(modelo__icontains=q)       |
                Q(n_serie__icontains=q)      |
                Q(unive__icontains=q)        |
                Q(bdo__icontains=q)          |
                Q(netbios__icontains=q)      |
                Q(ubicacion__icontains=q)    |
                Q(creado_por__first_name__icontains=q) |
                Q(creado_por__last_name__icontains=q)
            )

        # 2) Filtro por estado
        estado = self.request.GET.get('estado', '')
        if estado in ('Operativo', 'Dañado', 'Mala'):
            qs = qs.filter(estado=estado)

        return qs.order_by('n_serie')

    def get_context_data(self, **kwargs):
        # No llamamos a filtrar_y_paginar, sino al ListView + Paginator manual
        context = super().get_context_data(**kwargs)

        # Ya tenemos en context['page_obj'] el page_obj creado por ListView,
        # pero vamos a sobreescribirlo para forzar el queryset filtrado:
        qs = self.get_queryset()
        paginator = Paginator(qs, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context.update({
            'page_obj': page_obj,
            'paginator': paginator,
            'add_url': 'add_tablet',
            'nombre_activo': 'Tablet',
        })
        return context

@add_group_name_to_context
class AddTabletView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Vista para agregar nueva Tablet"""
    model = Tablet
    form_class = TabletForm # Necesitaremos crear este formulario
    template_name = 'modulos/add_tablet.html' # Necesitaremos crear esta plantilla
    success_url = reverse_lazy('tablet_list') # Necesitaremos crear esta URL

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name_singular'] = 'Tablet'
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                n_serie = form.cleaned_data.get('n_serie')
                unive = form.cleaned_data.get('unive')
                netbios = form.cleaned_data.get('netbios')

                if Tablet.objects.filter(n_serie=n_serie).exists():
                    form.add_error('n_serie', 'Este número de serie ya está registrado.')
                    return self.form_invalid(form)
                if unive != "0" and Tablet.objects.filter(unive=unive).exists():
                    form.add_error('unive', 'Este código UNIVE ya está registrado.')
                    return self.form_invalid(form)

                if netbios and Tablet.objects.filter(netbios=netbios).exists():
                    form.add_error('netbios', 'Este NetBIOS ya está registrado.')
                    return self.form_invalid(form)

                tablet = form.save(commit=False)
                tablet.creado_por = self.request.user
                tablet.save()

                # Notificación por correo
                try:
                    enviar_notificacion_asunto(
                        asunto="Nueva Tablet Registrada",
                        mensaje=f"""
                        El usuario {self.request.user.get_full_name()} ha agregado una nueva tablet al sistema.

                        Activo: {tablet.activo}
                        Marca: {tablet.marca}
                        Modelo: {tablet.modelo}
                        Número de Serie: {tablet.n_serie}
                        UNIVE: {tablet.unive}
                        NetBIOS: {tablet.netbios}
                        Ubicación: {tablet.ubicacion}
                        """,
                        destinatarios=settings.EMAIL_RECIPIENTS
                    )
                except Exception as e:
                    messages.error(self.request, f'Error al enviar el correo: {str(e)}')

                messages.success(self.request, 'Tablet agregada exitosamente')
                return super().form_valid(form)

        except IntegrityError as e:
            messages.error(self.request, 'Error de integridad de datos al guardar la tablet.')
            return self.form_invalid(form)

        except Exception as e:
            messages.error(self.request, f'Error inesperado al agregar la tablet: {str(e)}')
            return self.form_invalid(form)

@add_group_name_to_context
class EditTabletView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Vista para editar Tablet existente"""
    model = Tablet
    form_class = TabletForm
    template_name = 'modulos/edit_tablet.html' # Necesitaremos crear esta plantilla
    success_url = reverse_lazy('tablet_list')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name_singular'] = 'Tablet'
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                tablet = form.save(commit=False)
                original_tablet = Tablet.objects.get(pk=tablet.pk)
                changed_fields = []
                for field in form.changed_data:
                    old_value = getattr(original_tablet, field)
                    new_value = form.cleaned_data[field]
                    if old_value != new_value:
                        HistorialCambios.objects.create(
                            modelo=tablet.__class__.__name__,
                            objeto_id=tablet.pk,
                            usuario=self.request.user,
                            campo_modificado=field,
                            valor_anterior=str(old_value),
                            valor_nuevo=str(new_value)
                        )
                        changed_fields.append(f"{field}: de '{old_value}' a '{new_value}'")
                
                tablet.save()

                if changed_fields:
                    accion = "Tablet Editada"
                    nombre_completo_usuario = self.request.user.get_full_name()
                    campos_modificados_str = "\n".join(changed_fields)
                    mensaje = f"""
                    El usuario {nombre_completo_usuario} ha editado una tablet en el sistema.

                    Acción: {accion}
                    Activo: {tablet.activo}
                    Número de Serie: {tablet.n_serie}
                    Cambios realizados:
                    {campos_modificados_str}
                    """
                    try:
                        enviar_notificacion_asunto(
                            asunto="Tablet Editada en el Sistema",
                            mensaje=mensaje,
                            destinatarios=settings.EMAIL_RECIPIENTS
                        )
                    except Exception as e:
                        messages.error(self.request, f'Error al enviar el correo de notificación: {str(e)}')
                
                messages.success(self.request, 'Tablet actualizada exitosamente')
                return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, 'Error de integridad de datos al actualizar la tablet.')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'Error inesperado al actualizar la tablet: {str(e)}')
            return self.form_invalid(form)

@add_group_name_to_context
class TabletDetailView(LoginRequiredMixin, DetailView):
    model = Tablet
    template_name = 'modulos/detalle_tablet.html' # Necesitaremos crear esta plantilla
    context_object_name = 'tablet'

# -------- VISTAS DE HISTORIAL DE CAMBIOS --------

from django.utils.timezone import now
from django.utils.timezone import make_aware
from datetime import datetime
from django.views.generic.list import ListView
from .models import HistorialCambios
from django.contrib.auth.mixins import LoginRequiredMixin


@add_group_name_to_context
class HistorialCambiosView(LoginRequiredMixin, ListView):
    model = HistorialCambios
    template_name = 'historial_cambios.html'
    context_object_name = 'historial'
    paginate_by = 20
    
    def get_queryset(self):
        """
        Obtiene y configura el queryset de historial de cambios
        Aplica filtros por usuario, rango de fechas, modelo y ID de objeto si se proporcionan
        """
        queryset = super().get_queryset()
        usuario = self.request.GET.get('usuario', 'todos')
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')
        model_name = self.kwargs.get('model_name')
        object_id = self.kwargs.get('object_id')

        if usuario != 'todos':
            queryset = queryset.filter(usuario__username=usuario)

        if fecha_inicio:
            try:
                fecha_inicio_dt = make_aware(datetime.strptime(fecha_inicio, '%Y-%m-%d'))
                queryset = queryset.filter(fecha_modificacion__gte=fecha_inicio_dt)
            except ValueError:
                pass

        if fecha_fin:
            try:
                fecha_fin_dt = make_aware(datetime.strptime(fecha_fin, '%Y-%m-%d')).replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(fecha_modificacion__lte=fecha_fin_dt)
            except ValueError:
                pass

        if model_name and object_id:
            queryset = queryset.filter(modelo=model_name, objeto_id=object_id)

        # Verifica si el queryset está vacío
        if not queryset.exists():
            self.paginate_by = None  # Evita paginar si no hay resultados

        # Normalizar la capitalización del nombre del modelo para la visualización
        queryset = queryset.annotate(
            modelo_display=Case(
                When(modelo='allinoneadmins', then=Value('AllInOneAdmins')),
                default=F('modelo'),
                output_field=CharField()
            )
        )

        return queryset.order_by('-fecha_modificacion', '-pk')


    def get_context_data(self, **kwargs):
        """
        Agrega información adicional al contexto para el template.
        """
        context = super().get_context_data(**kwargs)
        context['usuario_seleccionado'] = self.request.GET.get('usuario', 'todos')
        context['fecha_inicio'] = self.request.GET.get('fecha_inicio', '')
        context['fecha_fin'] = self.request.GET.get('fecha_fin', '')
        context['usuarios'] = self.get_usuarios_disponibles()
        return context

    def get_usuarios_disponibles(self):
        """
        Retorna todos los usuarios activos disponibles para el filtro.
        """
        from django.contrib.auth.models import User
        return User.objects.filter(is_active=True)




@add_group_name_to_context
class AllInOneDetailView(DetailView):
    model = AllInOne
    template_name = 'modulos/detalle_all_in_one.html'
    context_object_name = 'allinone'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtener el historial de cambios para este objeto específico
        content_type = ContentType.objects.get_for_model(self.object)
        historial = HistorialCambios.objects.filter(
            objeto_id=self.object.pk,
            modelo=content_type.model
        ).order_by('-fecha_modificacion')
        context['historial_cambios'] = historial
        return context

@add_group_name_to_context
class AllInOneAdminDetailView(DetailView):
    model = AllInOneAdmins
    template_name = 'modulos/detalle_all_in_one_adm.html'
    context_object_name = 'allinoneadmins'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtener el historial de cambios para este objeto específico
        # Obtener el ContentType del modelo AllInOneAdmins
        content_type = ContentType.objects.get_for_model(AllInOneAdmins)
        
        historial = HistorialCambios.objects.filter(
            objeto_id=self.object.pk,
            modelo=content_type.model # Usar el nombre del modelo en minúsculas
        ).order_by('-fecha_modificacion', '-pk')
        context['historial_cambios'] = historial
        return context

@add_group_name_to_context
class NotebookDetailView(DetailView):
    model = Notebook
    template_name = 'modulos/detalle_notebook.html'
    context_object_name = 'notebook'
@add_group_name_to_context
class MiniPCDetailView(LoginRequiredMixin, DetailView):
    model = MiniPC
    template_name = 'modulos/detalle_mini_pc.html'
    context_object_name = 'minipc'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        minipc = self.object

        # Según cómo guardas 'modelo' en HistorialCambios:
        # - Si usas Clase.__name__ -> 'MiniPC'
        # - Si usas _meta.model_name -> 'minipc'
        context['historial_cambios'] = (
            HistorialCambios.objects.filter(
                Q(modelo=MiniPC.__name__) | Q(modelo=MiniPC._meta.model_name),
                objeto_id=minipc.pk
            ).order_by('-fecha_modificacion')
        )
        return context

@add_group_name_to_context
class ProyectorDetailView(LoginRequiredMixin, DetailView):
    model = Proyectores             # <-- si tu clase es Proyectores, cámbiala aquí
    template_name = 'modulos/detalle_proyector.html'
    context_object_name = 'proyector'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.object

        # Tu HistorialCambios no tiene content_type ni fecha_cambio
        context['historial_cambios'] = (
            HistorialCambios.objects
            .filter(
                Q(modelo=self.model.__name__) | Q(modelo=self.model._meta.model_name),
                objeto_id=obj.pk
            )
            .order_by('-fecha_modificacion')
        )

        # Para el botón "Volver al listado"
        context['list_url_name'] = 'proyectores'
        return context

class BodegaADRDetailView(DetailView):
    model = BodegaADR
    template_name = 'modulos/detalle_bodegaadr.html'   # ruta del template
    context_object_name = 'bodega'                       # cómo se llamará en el template
@add_group_name_to_context
class AzoteaDetailView(DetailView):
    model = Azotea
    template_name = 'modulos/detalle_azotea.html'
    context_object_name = 'azotea'










# -------- VISTA PARA CARGAR ARCHIVOS EXCEL --------

logger = logging.getLogger(__name__)

@add_group_name_to_context
class UploadExcelAllInOneView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'upload_excel_allinone.html'
    success_url = reverse_lazy('all_in_one')

    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('excel_file')
        logger.debug(f"Archivo subido: {excel_file}")
        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            return redirect('upload_excel_allinone')

        try:
            # Leer el archivo Excel
            df = pd.read_excel(excel_file)
            logger.debug(f"DataFrame leído correctamente con {len(df)} filas")

            # Validar que el archivo tenga las columnas esperadas
            required_columns = ['Estado', 'Activo', 'Marca', 'Modelo', 'N_serie', 'Unive', 'Bdo', 'Netbios', 'Fecha_creacion']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(self.request, f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                return self.form_invalid(form)

            # Rellenar valores nulos con "Operativo" para la columna "Estado"
            if 'Estado' in df.columns:
                df['Estado'] = df['Estado'].fillna(value='Operativo')

            # Procesar y crear registros AllInOne
            registros_existentes = 0
            nuevos_registros = 0

            for _, row in df.iterrows():
                if not AllInOne.objects.filter(
                    n_serie=row.get('N_serie', ''),
                    unive=row.get('Unive', ''),
                    bdo=row.get('Bdo', '')
                ).exists():
                    # Si no existe, crear un nuevo registro
                    AllInOne.objects.create(
                        estado=row['Estado'],
                        activo=row.get('Activo', 'All In One'),
                        marca=row.get('Marca', 'Seleccione'),
                        modelo=row.get('Modelo', ''),
                        n_serie=row.get('N_serie', ''),
                        unive=row.get('Unive', ''),
                        bdo=row.get('Bdo', ''),
                        netbios=row.get('Netbios', ''),
                        ubicacion=row.get('Ubicacion', 'Seleccione'),
                        creado_por=self.request.user,
                        fecha_creacion=row.get('Fecha_creacion', pd.Timestamp.now()),
                    )
                    nuevos_registros += 1
                    logger.debug(f"Registro creado para la fila {row.name + 2}")
                else:
                    # Contar registros existentes
                    registros_existentes += 1

            if nuevos_registros > 0:
                messages.success(self.request, f"El archivo Excel fue procesado exitosamente. {nuevos_registros} registros nuevos añadidos y {registros_existentes} ya existían.")
            else:
                messages.info(self.request, f"Todos los registros del archivo ya existen en la base de datos. {registros_existentes} registros encontrados, 0 nuevos añadidos.")

            # Enviar notificación por correo
            self.enviar_notificacion_correo(excel_file, nuevos_registros)

            return super().form_valid(form)

        except pd.errors.EmptyDataError:
            messages.error(self.request, "El archivo Excel está vacío. Por favor, sube un archivo válido.")
            return self.form_invalid(form)

        except Exception as e:
            # Captura del error general y mostrarlo en la plantilla
            logger.error(f"Error inesperado al procesar el archivo Excel: {str(e)}")
            messages.error(self.request, f"Hubo un error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def enviar_notificacion_correo(self, excel_file, nuevos_registros):
        """Envía un correo electrónico con la notificación de carga masiva"""
        try:
            accion = "Carga de datos masivos"
            modelo = "AllInOne"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Registros nuevos añadidos: {nuevos_registros}
            """

            email = EmailMessage(
                subject="Carga de Datos Masivos en AllInOne",
                body=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=settings.EMAIL_RECIPIENTS
            )

            # Adjuntar el archivo Excel subido
            excel_file.seek(0)  # Asegurarse de que el archivo esté en el inicio
            email.attach(excel_file.name, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            email.send()
            logger.info("Correo de notificación enviado exitosamente.")
        except Exception as e:
            logger.error(f"Error al enviar correo de notificación: {str(e)}")
        



logger = logging.getLogger(__name__)

@add_group_name_to_context
class UploadExcelNotebookView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'upload_excel_notebook.html'
    success_url = reverse_lazy('notebooks')

    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']

    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('excel_file')
        logger.debug(f"Archivo subido: {excel_file}")

        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            return redirect('upload_excel_notebook')

        try:
            # Leer el archivo Excel
            df = pd.read_excel(excel_file)
            logger.debug(f"DataFrame leído correctamente con {len(df)} filas")

            # Validar que el archivo tenga las columnas esperadas
            # Columnas del Excel del usuario: Activo, Modelo, N_serie, Unive, Bdo, Estado, Creado_por, Fecha_creacion, Fecha_modificacion, Marca, Ubicacion, Netbios
            required_columns = ['Estado', 'Activo', 'Marca', 'Modelo', 'N_serie', 'Unive', 'Bdo', 'Netbios', 'Fecha_creacion', 'Ubicación']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(self.request, f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                return self.form_invalid(form)

            # Rellenar valores nulos con valores predeterminados
            df['Estado'] = df['Estado'].fillna(value='Operativo')
            # El campo 'Asignado_a' no está en el Excel del usuario, se asignará un valor por defecto al crear.
            
            # Convertir la columna 'Bdo' a numérica, reemplazando 'nan' con None
            # Esto es crucial para DecimalField que no acepta 'nan'
            df['Bdo'] = pd.to_numeric(df['Bdo'], errors='coerce').replace({np.nan: None})

            # Procesar y crear registros Notebook, evitando duplicados
            registros_existentes = 0
            nuevos_registros = 0

            for _, row in df.iterrows():
                # Verificar si el registro ya existe basándonos en los identificadores clave
                # Usamos .get con '' para n_serie y unive para manejar posibles NaN que pandas podría leer como float
                n_serie_excel = str(row.get('N_serie', '')) if pd.notna(row.get('N_serie')) else None
                unive_excel = str(row.get('Unive', '')) if pd.notna(row.get('Unive')) else None
                bdo_excel = row.get('Bdo') # Ya convertido a None si es NaN

                if not Notebook.objects.filter(
                    n_serie=n_serie_excel,
                    unive=unive_excel,
                    bdo=bdo_excel
                ).exists():
                    # Si no existe, crear un nuevo registro
                    # Convertir Fecha_creacion a datetime si es necesario
                    fecha_creacion_excel = row.get('Fecha_creacion') # Corregido a Fecha_creacion con tilde
                    if pd.notna(fecha_creacion_excel):
                        try:
                            # Intentar convertir a datetime si es string u otro formato
                            fecha_creacion_dt = pd.to_datetime(fecha_creacion_excel)
                            # Asegurarse de que es timezone-aware si settings.USE_TZ es True
                            if settings.USE_TZ:
                                fecha_creacion_dt = make_aware(fecha_creacion_dt) if timezone.is_naive(fecha_creacion_dt) else fecha_creacion_dt
                        except ValueError:
                            # Si falla la conversión, usar la fecha actual o manejar el error
                            fecha_creacion_dt = timezone.now() if settings.USE_TZ else datetime.now()
                            logger.warning(f"No se pudo convertir Fecha_creacion '{fecha_creacion_excel}' para la fila {row.name + 2}. Usando fecha actual.")
                    else:
                        fecha_creacion_dt = timezone.now() if settings.USE_TZ else datetime.now()

                    Notebook.objects.create(
                        estado=row['Estado'],
                        activo=row.get('Activo', 'Notebook'), # Columna 'Activo' del Excel
                        marca=row.get('Marca', 'Seleccione'),
                        modelo=row.get('Modelo', ''),
                        n_serie=n_serie_excel,
                        unive=unive_excel,
                        bdo=bdo_excel,
                        netbios=str(row.get('Netbios', '')) if pd.notna(row.get('Netbios')) else None,
                        ubicacion=row.get('Ubicación', 'Seleccione'), # Corregido a Ubicacion con tilde
                        asignado_a='No asignado',  # Valor por defecto ya que no está en el Excel
                        creado_por=self.request.user,  # Asignar el usuario autenticado
                        fecha_creacion=fecha_creacion_dt, # Usar la fecha del Excel o la actual
                    )
                    nuevos_registros += 1
                    logger.debug(f"Registro creado para la fila {row.name + 2} with creado_por: {self.request.user}")
                    logger.debug(f"Registro creado para la fila {row.name + 2}")
                else:
                    # Contar registros existentes
                    registros_existentes += 1

            # Mensajes de éxito o información dependiendo de los resultados
            if nuevos_registros > 0:
                messages.success(self.request, f"El archivo Excel fue procesado exitosamente. {nuevos_registros} registros nuevos añadidos y {registros_existentes} ya existían.")
            else:
                messages.info(self.request, f"Todos los registros del archivo ya existen en la base de datos. {registros_existentes} registros encontrados, 0 nuevos añadidos.")

            # Enviar notificación por correo
            self.enviar_notificacion_correo(excel_file, nuevos_registros)

            return super().form_valid(form)

        except pd.errors.EmptyDataError:
            messages.error(self.request, "El archivo Excel está vacío. Por favor, sube un archivo válido.")
            return self.form_invalid(form)

        except Exception as e:
            # Captura del error general y mostrarlo en la plantilla
            logger.error(f"Error inesperado al procesar el archivo Excel: {str(e)}")
            messages.error(self.request, f"Hubo un error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def enviar_notificacion_correo(self, excel_file, nuevos_registros):
        """Envía un correo electrónico con la notificación de carga masiva"""
        try:
            accion = "Carga de datos masivos"
            modelo = "Notebook"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Registros nuevos añadidos: {nuevos_registros}
            """

            email = EmailMessage(
                subject="Carga de Datos Masivos en Notebook",
                body=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=settings.EMAIL_RECIPIENTS
            )

            # Adjuntar el archivo Excel subido
            excel_file.seek(0)  # Asegurarse de que el archivo esté en el inicio
            email.attach(excel_file.name, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            email.send()
            logger.info("Correo de notificación enviado exitosamente.")
        except Exception as e:
            logger.error(f"Error al enviar correo de notificación: {str(e)}")


logger = logging.getLogger(__name__)

@add_group_name_to_context
class UploadExcelMiniPCView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'upload_excel_minipc.html'
    success_url = reverse_lazy('mini_pc')
    
    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']
    
    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('excel_file')
        logger.debug(f"Archivo subido: {excel_file}")

        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            return redirect('upload_excel_minipc')

        try:
            # Leer el archivo Excel
            df = pd.read_excel(excel_file)
            logger.debug(f"DataFrame leído correctamente con {len(df)} filas")

            # Validar que el archivo tenga las columnas esperadas
            required_columns = ['Estado', 'Activo', 'Marca', 'Modelo', 'N_serie', 'Unive', 'Bdo', 'Netbios', 'Fecha_creacion']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(self.request, f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                return self.form_invalid(form)

            # Rellenar valores nulos con valores predeterminados
            df['Estado'] = df['Estado'].fillna(value='Operativo')
            df['Activo'] = df['Activo'].fillna(value='Mini PC')


            # Procesar y crear registros MiniPC, evitando duplicados
            registros_existentes = 0
            nuevos_registros = 0

            for _, row in df.iterrows():
                # Obtener valores y manejar NaN/vacío para campos que podrían ser numéricos
                unive_value = row.get('Unive', None)
                bdo_value = row.get('Bdo', None)

                # Convertir NaN o cadena vacía a un valor por defecto (0 para 'unive' ya que no puede ser null)
                if pd.isna(unive_value) or (isinstance(unive_value, str) and unive_value.strip() == ''):
                    unive_value = 0 # Usar 0 como valor por defecto para 'unive'
                # Convertir NaN o cadena vacía a None para campos que esperan números nulos (si 'bdo' permite null)
                if pd.isna(bdo_value) or (isinstance(bdo_value, str) and bdo_value.strip() == ''):
                    bdo_value = None # Mantener None para 'bdo' si permite null

                # Verificar si el registro ya existe basándonos en los identificadores clave
                # Usar los valores ya procesados (0 para 'unive', None si es NaN/vacío para 'bdo')
                if not MiniPC.objects.filter(
                    n_serie=row.get('N_serie', ''),
                    unive=unive_value,
                    bdo=bdo_value
                ).exists():
                    # Si no existe, crear un nuevo registro
                    MiniPC.objects.create(
                        estado=row['Estado'],
                        activo=row.get('Activo', 'MiniPC'),
                        marca=row.get('Marca', 'Seleccione'),
                        modelo=row.get('Modelo', ''),
                        n_serie=row.get('N_serie', ''),
                        unive=unive_value,
                        bdo=bdo_value,
                        netbios=row.get('Netbios', ''),
                        ubicacion=row.get('Ubicacion', 'Seleccione'),
                        creado_por=self.request.user,
                        fecha_creacion=row.get('Fecha_creacion', pd.Timestamp.now()),
                    )
                    nuevos_registros += 1
                    logger.debug(f"Registro creado para la fila {row.name + 2}")
                else:
                    # Contar registros existentes
                    registros_existentes += 1

            # Mensajes de éxito o información dependiendo de los resultados
            if nuevos_registros > 0:
                messages.success(self.request, f"El archivo Excel fue procesado exitosamente. {nuevos_registros} registros nuevos añadidos y {registros_existentes} ya existían.")
            else:
                messages.info(self.request, f"Todos los registros del archivo ya existen en la base de datos. {registros_existentes} registros encontrados, 0 nuevos añadidos.")

            # Enviar notificación por correo
            self.enviar_notificacion_correo(excel_file, nuevos_registros)

            return super().form_valid(form)

        except pd.errors.EmptyDataError:
            messages.error(self.request, "El archivo Excel está vacío. Por favor, sube un archivo válido.")
            return self.form_invalid(form)

        except Exception as e:
            # Captura del error general y mostrarlo en la plantilla
            logger.error(f"Error inesperado al procesar el archivo Excel: {str(e)}")
            messages.error(self.request, f"Hubo un error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def enviar_notificacion_correo(self, excel_file, nuevos_registros):
        """Envía un correo electrónico con la notificación de carga masiva"""
        try:
            accion = "Carga de datos masivos"
            modelo = "MiniPC"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Registros nuevos añadidos: {nuevos_registros}
            """

            email = EmailMessage(
                subject="Carga de Datos Masivos en MiniPC",
                body=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=settings.EMAIL_RECIPIENTS
            )

            # Adjuntar el archivo Excel subido
            excel_file.seek(0)  # Asegurarse de que el archivo esté en el inicio
            email.attach(excel_file.name, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            email.send()
            logger.info("Correo de notificación enviado exitosamente.")
        except Exception as e:
            logger.error(f"Error al enviar correo de notificación: {str(e)}")



logger = logging.getLogger(__name__)

@add_group_name_to_context
class UploadExcelProyectorView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'upload_excel_proyector.html'
    success_url = reverse_lazy('proyectores')
    
    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']
    
    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('excel_file')
        logger.debug(f"Archivo subido: {excel_file}")

        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            return redirect('upload_excel_proyector')

        try:
            # Leer el archivo Excel
            df = pd.read_excel(excel_file)
            logger.debug(f"DataFrame leído correctamente con {len(df)} filas")

            # Validar que el archivo tenga las columnas esperadas
            required_columns = ['Estado', 'Activo', 'Marca', 'Modelo', 'N_serie', 'Unive', 'Bdo', 'Fecha_creacion', 'Netbios']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(self.request, f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                return self.form_invalid(form)

            # Rellenar valores nulos con valores predeterminados donde corresponda
            df['Estado'] = df['Estado'].fillna(value='Operativo')
            df['Marca'] = df['Marca'].fillna(value='Seleccione')
            df['Ubicacion'] = df['Ubicacion'].fillna(value='Seleccione')

            # Reemplazar `nan` con `None` para los campos UNIVE y NETBIOS
            df['Unive'] = df['Unive'].replace({pd.NA: None, 'nan': None}).fillna(value='')
            df['Netbios'] = df['Netbios'].replace({pd.NA: None, 'nan': None}).fillna(value='')
            # Convertir NaN en Bdo a 0 y asegurar que sea de tipo numérico (entero en este caso)
            df['Bdo'] = pd.to_numeric(df['Bdo'], errors='coerce').fillna(value=0).astype(int)
            df['Activo'] = df['Activo'].fillna(value='Proyector')


            # Procesar y crear registros Proyector, evitando duplicados
            registros_existentes = 0
            nuevos_registros = 0

            for _, row in df.iterrows():
                # Verificar si el registro ya existe basándonos en los identificadores clave
                if not Proyectores.objects.filter(
                    n_serie=row.get('N_serie', ''),
                    unive=row.get('Unive', ''),
                    bdo=row.get('Bdo', 0)
                ).exists():
                    # Si no existe, crear un nuevo registro
                    Proyectores.objects.create(
                        estado=row['Estado'],
                        activo=row.get('Activo', 'Proyector'),
                        marca=row['Marca'],
                        modelo=row.get('Modelo', ''),
                        n_serie=row.get('N_serie', ''),
                        unive=row.get('Unive', ''),
                        bdo=row.get('Bdo', 0),
                        netbios=row.get('Netbios', ''),
                        ubicacion=row.get('Ubicacion', 'Seleccione'),
                        creado_por=self.request.user,
                        fecha_creacion=row.get('Fecha_creacion', pd.Timestamp.now()),
                    )
                    nuevos_registros += 1
                    logger.debug(f"Registro creado para la fila {row.name + 2}")
                else:
                    # Contar registros existentes
                    registros_existentes += 1

            # Mensajes de éxito o información dependiendo de los resultados
            if nuevos_registros > 0:
                messages.success(self.request, f"El archivo Excel fue procesado exitosamente. {nuevos_registros} registros nuevos añadidos y {registros_existentes} ya existían.")
            else:
                messages.info(self.request, f"Todos los registros del archivo ya existen en la base de datos. {registros_existentes} registros encontrados, 0 nuevos añadidos.")

            # Enviar notificación por correo
            self.enviar_notificacion_correo(excel_file, nuevos_registros)

            return super().form_valid(form)

        except pd.errors.EmptyDataError:
            messages.error(self.request, "El archivo Excel está vacío. Por favor, sube un archivo válido.")
            return self.form_invalid(form)

        except Exception as e:
            # Captura del error general y mostrarlo en la plantilla
            logger.error(f"Error inesperado al procesar el archivo Excel: {str(e)}")
            messages.error(self.request, f"Hubo un error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def enviar_notificacion_correo(self, excel_file, nuevos_registros):
        """Envía un correo electrónico con la notificación de carga masiva"""
        try:
            accion = "Carga de datos masivos"
            modelo = "Proyector"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Registros nuevos añadidos: {nuevos_registros}
            """

            email = EmailMessage(
                subject="Carga de Datos Masivos en Proyector",
                body=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=settings.EMAIL_RECIPIENTS
            )

            # Adjuntar el archivo Excel subido
            excel_file.seek(0)  # Asegurarse de que el archivo esté en el inicio
            email.attach(excel_file.name, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            email.send()
            logger.info("Correo de notificación enviado exitosamente.")
        except Exception as e:
            logger.error(f"Error al enviar correo de notificación: {str(e)}")



logger = logging.getLogger(__name__)

@add_group_name_to_context
class UploadExcelBodegaADRView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'upload_excel_bodega_adr.html'
    success_url = reverse_lazy('bodega_adr')
    
    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']
    
    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        try:
            df = pd.read_excel(excel_file)

            # Validar que el archivo tenga las columnas esperadas
            required_columns = ['Activo', 'Marca', 'Modelo', 'N_serie', 'Unive', 'Bdo', 'Netbios', 'Ubicacion', 'Fecha_creacion', 'Estado']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(self.request, f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                return self.form_invalid(form)

            # Procesamiento y creación de registros para BodegaADR, evitando duplicados
            registros_existentes = 0
            nuevos_registros = 0

            for _, row in df.iterrows():
                # Obtener valores y manejar NaN/vacío para campos que podrían ser numéricos o tener valor por defecto
                unive_value = row.get('Unive', None)
                bdo_value = row.get('Bdo', None)
                ubicacion_value = row.get('Ubicacion', 'Bueno') # Usar 'Ubicacion' del Excel para el campo renombrado 'ubicacion', con default 'Bueno'

                # Convertir NaN o cadena vacía a un valor por defecto (0 para 'unive' ya que no puede ser null)
                if pd.isna(unive_value) or (isinstance(unive_value, str) and unive_value.strip() == ''):
                    unive_value = None # Usar None como valor por defecto para 'unive' ya que es CharField(null=True)
                # Convertir NaN o cadena vacía a None para campos que esperan números nulos (si 'bdo' permite null)
                if pd.isna(bdo_value) or (isinstance(bdo_value, str) and bdo_value.strip() == ''):
                    bdo_value = None # Mantener None para 'bdo' if it allows null
                # ubicacion_value ya tiene un default 'Bueno' de .get()

                # Verificar si el registro ya existe basándonos en los identificadores clave
                # Usar los valores ya procesados
                if not BodegaADR.objects.filter(
                    n_serie=row.get('N_serie', ''),
                    unive=unive_value,
                    bdo=bdo_value
                ).exists():
                    # Si no existe, crear un nuevo registro
                    BodegaADR.objects.create(
                        ubicacion=ubicacion_value, # Usar el valor de 'Ubicacion' del Excel para el campo renombrado 'ubicacion'
                        estado=row.get('Estado', 'Activo'), # Mapear columna 'Estado' del Excel al campo 'estado' del modelo
                        activo=row.get('Activo', 'Bodega ADR'),
                        marca=row.get('Marca', 'Seleccione'),
                        modelo=row.get('Modelo', ''),
                        n_serie=row.get('N_serie', ''),
                        unive=unive_value,
                        bdo=bdo_value,
                        netbios=row.get('Netbios', ''),
                        creado_por=self.request.user,
                        fecha_creacion=row.get('Fecha_creacion', pd.Timestamp.now()),
                    )
                    nuevos_registros += 1
                    logger.debug(f"Registro creado para la fila {row.name + 2}")
                else:
                    # Contar registros existentes
                    registros_existentes += 1

            # Mensajes de éxito o información dependiendo de los resultados
            if nuevos_registros > 0:
                messages.success(self.request, f"El archivo Excel fue procesado exitosamente. {nuevos_registros} registros nuevos añadidos y {registros_existentes} ya existían.")
            else:
                messages.info(self.request, f"Todos los registros del archivo ya existen en la base de datos. {registros_existentes} registros encontrados, 0 nuevos añadidos.")

            # Enviar notificación por correo
            self.enviar_notificacion_correo(excel_file, nuevos_registros)

            return super().form_valid(form)

        except pd.errors.EmptyDataError:
            messages.error(self.request, "El archivo Excel está vacío. Por favor, sube un archivo válido.")
            return self.form_invalid(form)

        except Exception as e:
            # Captura del error general y mostrarlo en la plantilla
            logger.error(f"Error inesperado al procesar el archivo Excel: {str(e)}")
            messages.error(self.request, f"Hubo un error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def enviar_notificacion_correo(self, excel_file, nuevos_registros):
        """Envía un correo electrónico con la notificación de carga masiva"""
        try:
            accion = "Carga de datos masivos"
            modelo = "Bodega ADR"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Registros nuevos añadidos: {nuevos_registros}
            """

            email = EmailMessage(
                subject="Carga de Datos Masivos en Bodega ADR",
                body=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=settings.EMAIL_RECIPIENTS
            )

            # Adjuntar el archivo Excel subido
            excel_file.seek(0)  # Asegurarse de que el archivo esté en el inicio
            email.attach(excel_file.name, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            email.send()
            logger.info("Correo de notificación enviado exitosamente.")
        except Exception as e:
            logger.error(f"Error al enviar correo de notificación: {str(e)}")




import logging
from django.core.mail import EmailMessage
from django.conf import settings
from decimal import Decimal, InvalidOperation



logger = logging.getLogger(__name__)
@add_group_name_to_context
class UploadExcelAzoteaView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'upload_excel_azotea.html'
    success_url = reverse_lazy('azotea_adr')
    
    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']
    
    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        try:
            # Leer el archivo Excel
            df = pd.read_excel(excel_file)
            logger.debug(f"DataFrame leído correctamente con {len(df)} filas")

            # Validar que el archivo tenga las columnas esperadas
            # Excel columns: Id, Activo, Modelo, N_serie, Unive, Bdo, Ubicación, Creado, Fecha, Fecha_, Marca, Estado, Netbios
            required_columns = ['Activo', 'Modelo', 'N_serie', 'Unive', 'Bdo', 'Ubicación', 'Marca', 'Estado', 'Fecha_creacion'] # Netbios es opcional
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(self.request, f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                return self.form_invalid(form)

            # Procesamiento y creación de registros para Azotea, evitando duplicados
            registros_existentes = 0
            nuevos_registros = 0

            for _, row in df.iterrows():
                # Verificar si el registro ya existe basándonos en los identificadores clave
                # Preparar valores, convirtiendo "N/D" y NaN a None para campos CharField(null=True, blank=True)
                n_serie_val = row.get('N_serie')
                if isinstance(n_serie_val, str) and n_serie_val.upper() == 'N/D':
                    n_serie_val = None
                elif pd.isna(n_serie_val):
                    n_serie_val = None

                unive_val = row.get('Unive')
                if pd.isna(unive_val): unive_val = None
                
                bdo_excel_str = row.get('Bdo')
                bdo_val = None
                if pd.notna(bdo_excel_str):
                    s_bdo = str(bdo_excel_str)
                    # Quitar guiones, se podrían quitar otros caracteres si fuera necesario
                    cleaned_s_bdo = s_bdo.replace('-', '')
                    
                    if cleaned_s_bdo and cleaned_s_bdo.isdigit(): # Asegurar que solo sean dígitos
                        try:
                            bdo_val = Decimal(cleaned_s_bdo)
                        except InvalidOperation:
                            logger.warning(f"No se pudo convertir el valor BDO '{bdo_excel_str}' (limpio: '{cleaned_s_bdo}') a Decimal para la fila. Se guardará como None.")
                            bdo_val = None
                    elif cleaned_s_bdo: # Si después de limpiar no son solo dígitos
                        logger.warning(f"Valor BDO '{bdo_excel_str}' (limpio: '{cleaned_s_bdo}') contiene caracteres no numéricos. Se guardará como None.")
                        bdo_val = None
                    # Si cleaned_s_bdo está vacío (ej. BDO era solo '-'), bdo_val permanece None

                netbios_val = row.get('Netbios')
                if pd.isna(netbios_val): netbios_val = None
                
                # Verificar si el registro ya existe basándonos en los identificadores clave
                # Considerar que n_serie_val puede ser None
                filter_args = {}
                if n_serie_val is not None: # Solo filtrar por n_serie si no es None para evitar problemas con multiples None
                    filter_args['n_serie'] = n_serie_val
                # Añadir otros campos al filtro si son necesarios para unicidad y no son None
                # Por ejemplo, si unive y bdo juntos con n_serie (o modelo si n_serie es None) definen unicidad
                # Para este ejemplo, nos basaremos principalmente en n_serie si existe.
                # Si n_serie es None, la lógica de duplicados podría necesitar ser más robusta o aceptar duplicados.
                
                # Simplificamos la comprobación de existencia si n_serie es el principal identificador único cuando no es N/D
                exists = False
                if n_serie_val is not None:
                    exists = Azotea.objects.filter(n_serie=n_serie_val).exists()
                # Si n_serie es None, podrías querer comprobar por una combinación de otros campos o permitir la creación.
                # Por ahora, si n_serie es None, asumimos que no existe para crear uno nuevo (esto podría generar duplicados si n_serie no es el único identificador)

                if not exists:
                    # Si no existe, crear un nuevo registro
                    Azotea.objects.create(
                        ubicacion=row.get('Ubicación', ''), # Mapeado desde Ubicación del Excel y clave corregida
                        estado=row.get('Estado', 'Activo'),             # Mapeado desde Estado del Excel
                        activo=row.get('Activo'),
                        marca=row.get('Marca'),
                        modelo=row.get('Modelo'),
                        n_serie=n_serie_val,
                        unive=unive_val,
                        bdo=bdo_val,
                        netbios=netbios_val,
                        creado_por=self.request.user,
                        fecha_creacion=row.get('Fecha_creacion', pd.Timestamp.now()), # Asume que si 'Fecha_creacion' no está, se usa now()
                    )
                    nuevos_registros += 1
                else:
                    # Contar registros existentes
                    registros_existentes += 1

            # Mensajes de éxito o información dependiendo de los resultados
            if nuevos_registros > 0:
                messages.success(self.request, f"El archivo Excel fue procesado exitosamente. {nuevos_registros} registros nuevos añadidos y {registros_existentes} ya existían.")
            else:
                messages.info(self.request, f"Todos los registros del archivo ya existen en la base de datos. {registros_existentes} registros encontrados, 0 nuevos añadidos.")

            # Enviar notificación por correo
            self.enviar_notificacion_correo(excel_file, nuevos_registros)

            return super().form_valid(form)

        except pd.errors.EmptyDataError:
            messages.error(self.request, "El archivo Excel está vacío. Por favor, sube un archivo válido.")
            return self.form_invalid(form)

        except Exception as e:
            # Captura del error general y mostrarlo en la plantilla
            logger.error(f"Error inesperado al procesar el archivo Excel: {str(e)}")
            messages.error(self.request, f"Hubo un error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def enviar_notificacion_correo(self, excel_file, nuevos_registros):
        """Envía un correo electrónico con la notificación de carga masiva"""
        try:
            accion = "Carga de datos masivos"
            modelo = "Azotea ADR"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Registros nuevos añadidos: {nuevos_registros}
            """

            email = EmailMessage(
                subject="Carga de Datos Masivos en Azotea ADR",
                body=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=settings.EMAIL_RECIPIENTS
            )

            # Adjuntar el archivo Excel subido
            excel_file.seek(0)  # Asegurarse de que el archivo esté en el inicio
            email.attach(excel_file.name, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            email.send()
            logger.info("Correo de notificación enviado exitosamente.")
        except Exception as e:
            logger.error(f"Error al enviar correo de notificación: {str(e)}")

    def form_invalid(self, form):
        """Manejo de formulario inválido"""
        return super().form_invalid(form)





logger = logging.getLogger(__name__)

@add_group_name_to_context
class UploadExcelAllInOneAdmView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    form_class = UploadExcelForm
    template_name = 'upload_excel_allinoneadm.html'
    success_url = reverse_lazy('all_in_one_adm')
    
    def test_func(self):
        """Solo ADR puede editar"""
        return self.request.user.groups.first().name in ['ADR', 'Operadores ADR']
    
    def handle_no_permission(self):
        """Redirección si no tiene permisos"""
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('excel_file')
        logger.debug(f"Archivo subido: {excel_file}")

        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            return redirect('upload_excel_allinoneadm')

        try:
            # Leer el archivo Excel
            df = pd.read_excel(excel_file)
            logger.debug(f"DataFrame leído correctamente con {len(df)} filas")

            # Validar que el archivo tenga las columnas esperadas
            required_columns = ['Estado', 'Activo', 'Marca', 'Modelo', 'N_serie', 'Unive', 'Bdo', 'Netbios', 'Fecha_creacion']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(self.request, f"El archivo Excel no contiene las columnas necesarias: {', '.join(missing_columns)}")
                return self.form_invalid(form)

            # Rellenar valores nulos con "Operativo" para la columna "Estado"
            if 'Estado' in df.columns:
                df['Estado'] = df['Estado'].fillna(value='Operativo')

            # Procesar y crear registros AllInOneAdmins, evitando duplicados
            registros_existentes = 0
            nuevos_registros = 0

            for _, row in df.iterrows():
                # Obtener valores y manejar NaN/vacío para campos que podrían ser numéricos
                unive_value = row.get('Unive', None)
                bdo_value = row.get('Bdo', None)

                # Convertir NaN o cadena vacía a un valor por defecto (0 para 'unive' ya que no puede ser null)
                if pd.isna(unive_value) or (isinstance(unive_value, str) and unive_value.strip() == ''):
                    unive_value = 0 # Usar 0 como valor por defecto para 'unive'
                # Convertir NaN o cadena vacía a None para campos que esperan números nulos (si 'bdo' permite null)
                if pd.isna(bdo_value) or (isinstance(bdo_value, str) and bdo_value.strip() == ''):
                    bdo_value = None # Mantener None para 'bdo' si permite null

                # Verificar si el registro ya existe basándonos en los identificadores clave
                # Usar los valores ya procesados (0 para 'unive', None si es NaN/vacío para 'bdo')
                if not AllInOneAdmins.objects.filter(
                    n_serie=row.get('N_serie', ''),
                    unive=unive_value,
                    bdo=bdo_value
                ).exists():
                    # Si no existe, crear un nuevo registro
                    AllInOneAdmins.objects.create(
                        estado=row['Estado'],
                        activo=row.get('Activo', 'All In One Admin'),
                        marca=row.get('Marca', 'Seleccione'),
                        modelo=row.get('Modelo', ''),
                        n_serie=row.get('N_serie', ''),
                        unive=unive_value,
                        bdo=bdo_value,
                        netbios=row.get('Netbios', ''),
                        ubicacion=row.get('Ubicacion', 'Seleccione'),
                        creado_por=self.request.user,
                        fecha_creacion=row.get('Fecha_creacion', pd.Timestamp.now()),
                    )
                    nuevos_registros += 1
                    logger.debug(f"Registro creado para la fila {row.name + 2}")
                else:
                    # Contar registros existentes
                    registros_existentes += 1

            # Mensajes de éxito o información dependiendo de los resultados
            if nuevos_registros > 0:
                messages.success(self.request, f"El archivo Excel fue procesado exitosamente. {nuevos_registros} registros nuevos añadidos y {registros_existentes} ya existían.")
            else:
                messages.info(self.request, f"Todos los registros del archivo ya existen en la base de datos. {registros_existentes} registros encontrados, 0 nuevos añadidos.")

            # Enviar notificación por correo
            self.enviar_notificacion_correo(excel_file, nuevos_registros)

            return super().form_valid(form)

        except pd.errors.EmptyDataError:
            messages.error(self.request, "El archivo Excel está vacío. Por favor, sube un archivo válido.")
            return self.form_invalid(form)

        except Exception as e:
            # Captura del error general y mostrarlo en la plantilla
            logger.error(f"Error inesperado al procesar el archivo Excel: {str(e)}")
            messages.error(self.request, f"Hubo un error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def enviar_notificacion_correo(self, excel_file, nuevos_registros):
        """Envía un correo electrónico con la notificación de carga masiva"""
        try:
            accion = "Carga de datos masivos"
            modelo = "All In One Admin"
            mensaje = f"""
            El usuario {self.request.user.get_full_name()} ha realizado la siguiente acción:
            Acción: {accion}
            Modelo: {modelo}
            Registros nuevos añadidos: {nuevos_registros}
            """

            email = EmailMessage(
                subject="Carga de Datos Masivos en All In One Admin",
                body=mensaje,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=settings.EMAIL_RECIPIENTS
            )

            # Adjuntar el archivo Excel subido
            excel_file.seek(0)  # Asegurarse de que el archivo esté en el inicio
            email.attach(excel_file.name, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            email.send()
            logger.info("Correo de notificación enviado exitosamente.")
        except Exception as e:
            logger.error(f"Error al enviar correo de notificación: {str(e)}")

import re
from decimal import Decimal, InvalidOperation # Asegúrate que estas importaciones estén al inicio del archivo si no lo están ya.
def buscar_global(request):
    query_original_case_sensitive = request.GET.get("q", "").strip()
    if not query_original_case_sensitive:  # Si 'q' está vacío o no existe, intentar con 'search'
        query_original_case_sensitive = request.GET.get("search", "").strip()
    
    print(f"[DEBUG] buscar_global: query_original_case_sensitive = '{query_original_case_sensitive}'")  # DEBUG
    query_original = query_original_case_sensitive.lower()
    print(f"[DEBUG] buscar_global: query_original (lowercase) = '{query_original}'")  # DEBUG
    
    query_sin_prefijo_scanner = query_original
    if re.match(r"^[a-z]{3}", query_original):  # ej., "ser123" -> "123"
        query_sin_prefijo_scanner = query_original[3:]

    # Mapeo de modelos a sus nombres plurales y nombres de URL para detalles
    # El 'url_detalle_name' es crucial para el enlace 'Ver' en la plantilla.
    modelos_info = [
        {'modelo_class': Notebook, 'nombre_plural': 'Notebooks', 'url_detalle_name': 'detalle_notebook', 'modelo_name': 'Notebook'},
        {'modelo_class': MiniPC, 'nombre_plural': 'Mini PCs', 'url_detalle_name': 'detalle_minipc', 'modelo_name': 'MiniPC'},
        {'modelo_class': Proyectores, 'nombre_plural': 'Proyectores', 'url_detalle_name': 'detalle_proyector', 'modelo_name': 'Proyectores'},
        {'modelo_class': AllInOne, 'nombre_plural': 'All In Ones', 'url_detalle_name': 'detalle_allinone', 'modelo_name': 'AllInOne'},
        {'modelo_class': AllInOneAdmins, 'nombre_plural': 'All In Ones Administrativos', 'url_detalle_name': 'detalle_allinone_admin', 'modelo_name': 'AllInOneAdmins'},
        {'modelo_class': BodegaADR, 'nombre_plural': 'Equipos en Bodega ADR', 'url_detalle_name': 'detalle_bodegaadr', 'modelo_name': 'BodegaADR'},
        {'modelo_class': Azotea, 'nombre_plural': 'Equipos en Azotea', 'url_detalle_name': 'detalle_azotea', 'modelo_name': 'Azotea'},
        {'modelo_class': Tablet, 'nombre_plural': 'Tablets', 'url_detalle_name': 'detalle_tablet', 'modelo_name': 'Tablet'},
        {'modelo_class': Monitor, 'nombre_plural': 'Monitores', 'url_detalle_name': 'detalle_monitor', 'modelo_name': 'Monitor'},
        {'modelo_class': Audio, 'nombre_plural': 'Equipos de Audio', 'url_detalle_name': 'detalle_audio', 'modelo_name': 'Audio'},
    ]

    resultados_por_modelo = []
    total_resultados = 0

    if query_original_case_sensitive:  # Solo buscar si hay un término de búsqueda
        for info in modelos_info:
            modelo = info['modelo_class']
            q_expressions = Q()
            
            campos_texto_generales = ['activo', 'modelo', 'unive', 'estado', 'marca', 'netbios', 'ubicacion']
            # Para Notebook, también buscar en 'asignado_a'
            if info['modelo_name'] == 'Notebook':
                campos_texto_generales.append('asignado_a')

            for campo in campos_texto_generales:
                if hasattr(modelo, campo):
                    q_expressions |= Q(**{f"{campo}__icontains": query_original})
            
            if hasattr(modelo, 'bdo'):
                q_expressions |= Q(bdo__icontains=query_original)

            if hasattr(modelo, 'n_serie'):
                q_n_serie = Q(n_serie__icontains=query_original)
                if query_sin_prefijo_scanner != query_original and query_sin_prefijo_scanner:  # Asegurar que no sea vacío
                    q_n_serie |= Q(n_serie__icontains=query_sin_prefijo_scanner)
                q_expressions |= q_n_serie
                
            if hasattr(modelo, 'creado_por'):
                q_expressions |= (
                    Q(creado_por__first_name__icontains=query_original) |
                    Q(creado_por__last_name__icontains=query_original) |
                    Q(creado_por__username__icontains=query_original)
                )
            
            if q_expressions:  # Solo filtrar si se construyó alguna expresión Q
                try:
                    resultados_modelo = modelo.objects.select_related('creado_por').filter(q_expressions)
                except Exception:
                    resultados_modelo = modelo.objects.filter(q_expressions)

                if resultados_modelo.exists():
                    resultados_por_modelo.append({
                        'nombre_plural': info['nombre_plural'],
                        'resultados': list(resultados_modelo),  # Convertir a lista para la plantilla
                        'modelo_name': info['modelo_name']
                    })
                    total_resultados += resultados_modelo.count()

    
    # Las condiciones if/elif pass anteriores fueron eliminadas por redundancia.
    # La plantilla maneja 'resultados_por_modelo' vacío.

    group_name_for_template = None
    if request.user.is_authenticated and request.user.groups.exists():
        group_for_template = request.user.groups.first()
        group_name_for_template = group_for_template.name

    context = {
        'query': query_original_case_sensitive, # Mostrar el query original en la plantilla
        'resultados_por_modelo': resultados_por_modelo,
        'total_resultados': total_resultados,
        'group_name_singular': group_name_for_template, # Asegurar que group_name_singular esté en el contexto
    }
    return render(request, 'resultados_busqueda_global.html', context)


@login_required # Se mantiene login_required, se remueve @add_group_name_to_context
def detalle_activo_busqueda(request, model_name, pk):
    """
    Muestra los detalles de un activo específico y su último historial de cambios.
    """
    try:
        # Obtener el tipo de contenido (modelo)
        # Asegurarse de que model_name esté en minúsculas para la búsqueda de ContentType
        content_type = ContentType.objects.get(model=model_name.lower())
        modelo_clase = content_type.model_class()
        
        # Obtener la instancia del activo, cargando explícitamente el usuario relacionado
        activo_instance = get_object_or_404(modelo_clase.objects.select_related('creado_por'), pk=pk)
        
        # Obtener todo el historial de cambios para este activo
        # Comparar con el nombre del modelo en minúsculas o el verbose_name según cómo se guarde en HistorialCambios
        # Asumiendo que HistorialCambios.modelo guarda el verbose_name del modelo.
        # Considerar tanto el verbose_name del modelo como la posible inconsistencia "MiniPC"
        modelo_nombres = [modelo_clase._meta.verbose_name]
        # Añadir nombres inconsistentes si el modelo es MiniPC, AllInOneAdmins, Proyectores, Audio, BodegaADR o Azotea
        if modelo_clase._meta.model_name == 'minipc':
             modelo_nombres.append('MiniPC')
        elif modelo_clase._meta.model_name == 'allinoneadmins':
             modelo_nombres.append('allinoneadmins')
        elif modelo_clase._meta.model_name == 'proyectores':
             modelo_nombres.append('proyectores')
        elif modelo_clase._meta.model_name == 'audio':
             modelo_nombres.append('audio')
             modelo_nombres.append('Audio')
        elif modelo_clase._meta.model_name == 'bodegaadr':
             modelo_nombres.append('bodegaadr')
        elif modelo_clase._meta.model_name == 'azotea': # Verificar si es el modelo Azotea
             modelo_nombres.append('azotea') # Añadir la posible inconsistencia "azotea"


        historial_completo = HistorialCambios.objects.filter(
            Q(modelo__in=modelo_nombres), # Usar Q object para buscar en la lista de nombres
            objeto_id=activo_instance.pk
        ).order_by('-fecha_modificacion') # Obtener todo el historial ordenado por fecha


        # Obtener información del grupo y color para el contexto
        # Verificar si el usuario está autenticado antes de llamar a get_group_and_color
        group_id, group_name, group_name_singular, color = None, None, None, None
        if request.user.is_authenticated:
            group_id, group_name, group_name_singular, color = get_group_and_color(request.user)

        # Obtener el término de búsqueda original de los parámetros GET
        original_query = request.GET.get('q', '')

        context = {
            'activo': activo_instance,
            'historial_cambios': historial_completo, # Pasar el historial completo al contexto
            'model_name_display': modelo_clase._meta.verbose_name.capitalize(), # Para mostrar en la plantilla
            'nombre_plural_modelo': modelo_clase._meta.verbose_name_plural.capitalize(), # Para breadcrumbs o títulos
            'group_name': group_name,
            'group_name_singular': group_name_singular,
            'color': color,
            'original_query': original_query, # Pasar el término de búsqueda a la plantilla
        }
        return render(request, 'detalle_activo_busqueda.html', context)

    except ContentType.DoesNotExist:
        raise Http404(f"El modelo '{model_name}' no existe o no se pudo encontrar.")
    except modelo_clase.DoesNotExist: # Captura específica si el activo no se encuentra
        raise Http404(f"El activo de tipo '{model_name}' con ID '{pk}' no existe.")
    except Exception as e:
        # Loggear el error para depuración
        # logger.error(f"Error al obtener detalles del activo ({model_name}, {pk}): {e}", exc_info=True)
        messages.error(request, f"Error al obtener detalles del activo: {str(e)}")
        # Redirigir a una página anterior o a home
        return redirect(request.META.get('HTTP_REFERER', reverse('home')))


@add_group_name_to_context
class UploadExcelMonitorView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    """Vista para subir archivo Excel de Monitores"""
    form_class = UploadExcelForm
    template_name = 'upload_excel_monitor.html'
    success_url = reverse_lazy('monitor_list')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def form_valid(self, form):
        print("form_valid called for UploadExcelMonitorView") # Debug print
        excel_file = form.cleaned_data.get('excel_file')
        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            print("No file selected.") # Debug print
            return self.form_invalid(form)

        try:
            df = pd.read_excel(excel_file)
            print(f"DataFrame read successfully. Shape: {df.shape}") # Debug print
            print(f"DataFrame columns: {df.columns.tolist()}") # Debug print
            df.columns = df.columns.str.lower().str.replace(' ', '_')
            print(f"Normalized columns: {df.columns.tolist()}") # Debug print

            column_mapping = {
                'estado': 'estado',
                'activo': 'activo',
                'marca': 'marca',
                'modelo': 'modelo',
                'n_serie': 'n_serie',
                'unive': 'unive',
                'bdo': 'bdo',
                'ubicacion': 'ubicacion',
                'fecha_creacion': 'fecha_creacion',
                'fecha_modificacion': 'fecha_modificacion',
                'creado_por': 'creado_por',
                'asignado_a': 'asignado_a', # Añadir mapeo para la columna 'Asignado_a'
            }

            valid_columns = {col: field for col, field in column_mapping.items() if col in df.columns}
            print(f"Valid columns for mapping: {valid_columns}") # Debug print

            created_count = 0
            updated_count = 0
            errors = []

            with transaction.atomic():
                print("Starting atomic transaction.") # Debug print
                # Filtrar datos para incluir solo campos válidos del modelo Monitor
                monitor_fields = [f.name for f in Monitor._meta.concrete_fields]
                print(f"Monitor fields: {monitor_fields}") # Debug print

                for index, row in df.iterrows():
                    try:
                        print(f"Processing row {index + 2}: {row.to_dict()}") # Debug print
                        data = {field: row[col] if pd.notna(row[col]) else None for col, field in valid_columns.items()}

                        # Usar fecha de creación del Excel si existe, de lo contrario usar la actual
                        if 'fecha_creacion' not in data or pd.isna(data['fecha_creacion']):
                            data['fecha_creacion'] = timezone.now()
                        else:
                             # Asegurarse de que la fecha sea un objeto datetime con zona horaria
                             if isinstance(data['fecha_creacion'], datetime):
                                 if data['fecha_creacion'].tzinfo is None:
                                     data['fecha_creacion'] = make_aware(data['fecha_creacion'])
                             else:
                                 # Intentar convertir si no es datetime (puede ser un string, etc.)
                                 try:
                                     data['fecha_creacion'] = make_aware(datetime.strptime(str(data['fecha_creacion']), '%Y-%m-%d %H:%M:%S')) # Ajusta el formato si es necesario
                                 except:
                                     data['fecha_creacion'] = timezone.now() # Usar fecha actual si falla la conversión


                        # Usar fecha de modificación del Excel si existe, de lo contrario usar la actual
                        if 'fecha_modificacion' not in data or pd.isna(data['fecha_modificacion']):
                            data['fecha_modificacion'] = timezone.now()
                        else:
                             # Asegurarse de que la fecha sea un objeto datetime con zona horaria
                             if isinstance(data['fecha_modificacion'], datetime):
                                 if data['fecha_modificacion'].tzinfo is None:
                                     data['fecha_modificacion'] = make_aware(data['fecha_modificacion'])
                             else:
                                 # Intentar convertir si no es datetime (puede ser un string, etc.)
                                 try:
                                     data['fecha_modificacion'] = make_aware(datetime.strptime(str(data['fecha_modificacion']), '%Y-%m-%d %H:%M:%S')) # Ajusta el formato si es necesario
                                 except:
                                     data['fecha_modificacion'] = timezone.now() # Usar fecha actual si falla la conversión


                        # Intentar encontrar un usuario basado en el nombre del Excel en la columna 'creado_por'
                        creado_por_excel_nombre = data.get('creado_por', '')
                        linked_user = None
                        if creado_por_excel_nombre:
                            # Buscar por username, first_name o last_name
                            try:
                                linked_user = User.objects.get(
                                    Q(username__iexact=creado_por_excel_nombre) |
                                    Q(first_name__iexact=creado_por_excel_nombre) |
                                    Q(last_name__iexact=creado_por_excel_nombre) |
                                    Q(first_name__iexact=creado_por_excel_nombre.split(' ')[0], last_name__iexact=' '.join(creado_por_excel_nombre.split(' ')[1:])) # Buscar por nombre y apellido
                                )
                                print(f"Linked user found: {linked_user.username}") # Debug print
                            except User.DoesNotExist:
                                print(f"No user found for name: {creado_por_excel_nombre}") # Debug print
                                pass # No se encontró usuario, linked_user sigue siendo None
                            except Exception as user_lookup_error:
                                print(f"Error during user lookup for {creado_por_excel_nombre}: {user_lookup_error}") # Debug print
                                pass # Manejar otros posibles errores de búsqueda
                            
                        # Asignar el usuario actualmente logueado al campo creado_por
                        data['creado_por'] = self.request.user
                        
                        print(f"Prepared data for update_or_create: {data}") # Debug print
                        
                        if not data.get('n_serie'):
                            errors.append(f"Fila {index + 2}: Número de serie faltante.")
                            print(f"Error: Missing n_serie in row {index + 2}") # Debug print
                            continue
                        if not data.get('modelo'):
                             errors.append(f"Fila {index + 2}: Modelo faltante.")
                             print(f"Error: Missing modelo in row {index + 2}") # Debug print
                             continue
                         
                        filtered_data = {key: value for key, value in data.items() if key in monitor_fields}
                        print(f"Filtered data for update_or_create: {filtered_data}") # Debug print
                        
                        instance, created = Monitor.objects.update_or_create(
                            n_serie=filtered_data.get('n_serie'), # Usar n_serie del diccionario filtrado
                            defaults=filtered_data
                        )
                        if created:
                            created_count += 1
                            print(f"Created Monitor: {instance.n_serie}") # Debug print
                            # Registrar historial de creación
                            HistorialCambios.objects.create(
                                usuario=self.request.user,
                                modelo="Monitor",  # Mapea modelo_afectado a modelo
                                objeto_id=instance.id,
                                campo_modificado="Creación", # Mapea accion a campo_modificado
                                valor_nuevo=f"Monitor creado mediante carga masiva. N/Serie: {instance.n_serie}", # Mapea detalles a valor_nuevo
                                valor_anterior="" # Deja valor_anterior vacío para creación
                            )
                        else:
                            updated_count += 1
                            print(f"Updated Monitor: {instance.n_serie}") # Debug print
                            # Registrar historial de modificación
                            HistorialCambios.objects.create(
                                usuario=self.request.user,
                                modelo="Monitor", # Mapea modelo_afectado a modelo
                                objeto_id=instance.id,
                                campo_modificado="Actualización", # Mapea accion a campo_modificado
                                valor_nuevo=f"Monitor actualizado mediante carga masiva. N/Serie: {instance.n_serie}", # Mapea detalles a valor_nuevo
                                valor_anterior="" # Podrías añadir lógica para capturar valores anteriores si es necesario
                            )
                    except Exception as e:
                        errors.append(f"Fila {index + 2}: Error al procesar el registro - Tipo: {type(e).__name__}, Mensaje: {e}")
                        print(f"Exception processing row {index + 2}: Tipo: {type(e).__name__}, Mensaje: {e}") # Debug print

            print(f"Transaction finished. Created: {created_count}, Updated: {updated_count}, Errors: {len(errors)}") # Debug print

            if errors:
                for error in errors:
                    messages.error(self.request, error)
                messages.warning(self.request, f"Proceso completado con {created_count} creados, {updated_count} actualizados y {len(errors)} errores.")
            else:
                messages.success(self.request, f"Archivo Excel procesado exitosamente. {created_count} registros creados, {updated_count} actualizados.")

             # Enviar notificación por correo
            asunto = "Carga Masiva de Monitores Completada"
            # 1) Preparamos la cadena con los detalles de errores (sin usar f-string)
            detalle_errores = "\n".join(errors) if errors else "No se encontraron errores."

            # 2) Construimos el mensaje usando esa variable dentro del f-string
            mensaje = f"""
            Se ha completado la carga masiva de monitores mediante archivo Excel.

            Resumen del proceso:
            - Registros creados: {created_count}
            - Registros actualizados: {updated_count}
            - Errores encontrados: {len(errors)}

            Detalles de los errores (si los haya):
            {detalle_errores}
            """

            try:
                enviar_notificacion_asunto(
                    asunto=asunto,
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS,
                    adjunto=excel_file # Pasar el archivo Excel como adjunto
                )
                print("Correo de notificación de carga masiva de monitores enviado con archivo adjunto.")
            except Exception as e:
                print(f"Error al enviar el correo de notificación de carga masiva con adjunto: {str(e)}")
                messages.error(self.request, 'Error al enviar el correo de notificación de carga masiva con adjunto.')

            return super().form_valid(form)

        except FileNotFoundError:
            messages.error(self.request, "Error: Archivo no encontrado.")
            print("FileNotFoundError") # Debug print
            return self.form_invalid(form)
        except pd.errors.EmptyDataError:
            messages.error(self.request, "Error: El archivo Excel está vacío.")
            print("EmptyDataError") # Debug print
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f"Error al procesar el archivo Excel: {str(e)}")
            print(f"General Exception: {e}") # Debug print
            return self.form_invalid(form)

    def form_invalid(self, form):
        print("form_invalid called for UploadExcelMonitorView") # Debug print
        messages.error(self.request, 'Error al subir el archivo. Por favor, verifique el formulario.')
        return super().form_invalid(form)

# Add email notification method if needed, similar to other upload views


@add_group_name_to_context
class UploadExcelAudioView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    """Vista para subir archivo Excel de Audio"""
    form_class = UploadExcelForm
    template_name = 'upload_excel_audio.html'
    success_url = reverse_lazy('audio_list')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('excel_file')
        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            return self.form_invalid(form)

        try:
            df = pd.read_excel(excel_file)
            df.columns = df.columns.str.lower().str.replace(' ', '_')

            column_mapping = {
                'estado': 'estado',
                'activo': 'activo',
                'marca': 'marca',
                'modelo': 'modelo',
                'n_serie': 'n_serie',
                'unive': 'unive',
                'bdo': 'bdo',
                'ubicacion': 'ubicacion',
            }

            valid_columns = {col: field for col, field in column_mapping.items() if col in df.columns}

            created_count = 0
            updated_count = 0
            errors = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    data = {}
                    for col, field in valid_columns.items():
                        value = row[col]
                        if field == 'bdo':
                            if pd.notna(value):
                                try:
                                    data[field] = Decimal(str(value))
                                except InvalidOperation:
                                    data[field] = None # Asignar None si la conversión falla
                                    errors.append(f"Fila {index + 2}: El valor '{value}' en BDO no es un número válido.")
                            else:
                                data[field] = None
                        else:
                            data[field] = value if pd.notna(value) else None

                    # Asignar usuario actual y fecha de modificación si no están en el Excel
                    data['creado_por'] = self.request.user
                    data['fecha_modificacion'] = timezone.now()

                    print(f"DEBUG: Datos preparados para la fila {index + 2}: {data}")

                    n_serie_value = data.get('n_serie')

                    if n_serie_value:
                        # Si hay un número de serie, intentar actualizar o crear
                        try:
                            instance, created = Audio.objects.update_or_create(
                                n_serie=n_serie_value,
                                defaults=data
                            )
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        except Exception as e:
                            errors.append(f"Fila {index + 2}: Error al procesar el registro con N_serie '{n_serie_value}' - {str(e)}")
                            print(f"DEBUG: Error al procesar fila {index + 2} con N_serie '{n_serie_value}': {str(e)}")
                    else:
                        # Si no hay número de serie, crear un nuevo registro
                        try:
                            # Asegurarse de que 'marca' sea una cadena vacía si es None, ya que CharField no puede ser None
                            if 'marca' in data and data['marca'] is None:
                                data['marca'] = ''
                            # Eliminar n_serie de los datos si es None para evitar errores de campo vacío en la creación
                            if 'n_serie' in data:
                                del data['n_serie']
                            
                            instance = Audio.objects.create(**data)
                            created_count += 1
                        except Exception as e:
                            errors.append(f"Fila {index + 2}: Error al crear nuevo registro (N_serie vacío) - {str(e)}")
                            print(f"DEBUG: Error al crear nuevo registro en fila {index + 2} (N_serie vacío): {str(e)}")

            if errors:
                for error in errors:
                    messages.error(self.request, error)
                messages.warning(self.request, f"Proceso completado con {created_count} registros creados, {updated_count} actualizados y {len(errors)} errores.")
            else:
                messages.success(self.request, f"Archivo Excel procesado exitosamente. {created_count} registros creados, {updated_count} actualizados.")

           # Enviar notificación por correo
            asunto = "Carga Masiva de Equipos de Audio Completada"

            # 1) Construye la cadena de detalle de errores fuera del f-string
            detalle_errores = "\n".join(errors) if errors else "No se encontraron errores."

            # 2) Usa esa variable dentro del f-string (sin barras invertidas en las llaves)
            mensaje = f"""
            Se ha completado la carga masiva de equipos de audio mediante archivo Excel.

            Resumen del proceso:
            - Registros creados: {created_count}
            - Registros actualizados: {updated_count}
            - Errores encontrados: {len(errors)}

            Detalles de los errores (si los haya):
            {detalle_errores}
            """

            try:
                enviar_notificacion_asunto(
                    asunto=asunto,
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS,
                    adjunto=excel_file # Pasar el archivo Excel como adjunto
                )
                print("Correo de notificación de carga masiva de equipos de audio enviado con archivo adjunto.")
            except Exception as e:
                print(f"Error al enviar el correo de notificación de carga masiva de audio con adjunto: {str(e)}")
                messages.error(self.request, 'Error al enviar el correo de notificación de carga masiva de audio con adjunto.')

            return super().form_valid(form)

        except FileNotFoundError:
            messages.error(self.request, "Error: Archivo no encontrado.")
            return self.form_invalid(form)
        except pd.errors.EmptyDataError:
            messages.error(self.request, "Error: El archivo Excel está vacío.")
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f"Error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Error al subir el archivo. Por favor, verifique el formulario.')
        return super().form_invalid(form)

# Add email notification method if needed, similar to other upload views


@add_group_name_to_context
class UploadExcelTabletView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    """Vista para subir archivo Excel de Tablets"""
    form_class = UploadExcelForm
    template_name = 'upload_excel_tablet.html'
    success_url = reverse_lazy('tablet_list')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['ADR', 'Operadores ADR']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('excel_file')
        if not excel_file:
            messages.error(self.request, "No se ha seleccionado ningún archivo.")
            return self.form_invalid(form)

        try:
            df = pd.read_excel(excel_file)
            df.columns = df.columns.str.lower().str.replace(' ', '_')

            column_mapping = {
                'estado': 'estado',
                'activo': 'activo',
                'marca': 'marca',
                'modelo': 'modelo',
                'n_serie': 'n_serie',
                'unive': 'unive',
                'bdo': 'bdo',
                'netbios': 'netbios',
                'almacenamiento': 'almacenamiento',
                'ubicacion': 'ubicacion',
                'fecha_creacion': 'fecha_creacion',
            }

            valid_columns = {col: field for col, field in column_mapping.items() if col in df.columns}

            created_count = 0
            updated_count = 0
            errors = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    data = {}
                    for col, field in valid_columns.items():
                        data[field] = row[col] if pd.notna(row[col]) else None

                    # Asignar usuario actual y fecha de modificación si no están en el Excel
                    data['creado_por'] = self.request.user
                    data['fecha_modificacion'] = timezone.now()

                    if not data.get('n_serie'):
                        errors.append(f"Fila {index + 2}: Número de serie faltante.")
                        continue
                    if not data.get('modelo'):
                         errors.append(f"Fila {index + 2}: Modelo faltante.")
                         continue

                    try:
                        instance, created = Tablet.objects.update_or_create(
                            n_serie=data['n_serie'],
                            defaults=data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    except Exception as e:
                        errors.append(f"Fila {index + 2}: Error al procesar el registro - {str(e)}")

            if errors:
                for error in errors:
                    messages.error(self.request, error)
                messages.warning(self.request, f"Proceso completado con {created_count} creados, {updated_count} actualizados y {len(errors)} errores.")
            else:
                messages.success(self.request, f"Archivo Excel procesado exitosamente. {created_count} registros creados, {updated_count} actualizados.")

          # Enviar notificación por correo
            asunto = "Carga Masiva de Tablets Completada"

            # 1) Prepara la cadena de errores fuera del f-string
            detalle_errores = "\n".join(errors) if errors else "No se encontraron errores."

            # 2) Ahora construye el mensaje sin backslashes dentro de las llaves
            mensaje = f"""
            Se ha completado la carga masiva de tablets mediante archivo Excel.

            Resumen del proceso:
            - Registros creados: {created_count}
            - Registros actualizados: {updated_count}
            - Errores encontrados: {len(errors)}

            Detalles de los errores (si los haya):
            {detalle_errores}
            """
            try:
                enviar_notificacion_asunto(
                    asunto=asunto,
                    mensaje=mensaje,
                    destinatarios=settings.EMAIL_RECIPIENTS,
                    adjunto=excel_file # Pasar el archivo Excel como adjunto
                )
                print("Correo de notificación de carga masiva de tablets enviado con archivo adjunto.")
            except Exception as e:
                print(f"Error al enviar el correo de notificación de carga masiva de tablets con adjunto: {str(e)}")
                messages.error(self.request, 'Error al enviar el correo de notificación de carga masiva de tablets con adjunto.')

            return super().form_valid(form)

        except FileNotFoundError:
            messages.error(self.request, "Error: Archivo no encontrado.")
            return self.form_invalid(form)
        except pd.errors.EmptyDataError:
            messages.error(self.request, "Error: El archivo Excel está vacío.")
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f"Error al procesar el archivo Excel: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Error al subir el archivo. Por favor, verifique el formulario.')
        return super().form_invalid(form)

# Add email notification method if needed, similar to other upload views